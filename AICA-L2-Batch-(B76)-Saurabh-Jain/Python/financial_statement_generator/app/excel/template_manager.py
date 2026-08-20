"""Inspect and preserve the ICAI Division I Excel template without modifying it.

The original workbook is treated as read-only. Later generation must copy this
file and write values into the copy so formatting and formulas stay intact.
"""

from __future__ import annotations

import hashlib
import json
import shutil
from dataclasses import dataclass, field
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.formula import ArrayFormula, DataTableFormula
from openpyxl.worksheet.worksheet import Worksheet

from app.config import settings

EXPECTED_SHEET_NAMES: tuple[str, ...] = (
    "BS PnL",
    "Cash Flow",
    "Note 1-2",
    "Note 3 (Share Capital)",
    "NOTE (4-12)",
    "Note 12 (PPE)",
    "13.CWIP",
    "Note (13-20)",
    "Note 20-31",
    "EBP",
    "Ratio_working",
    "Other Notes",
    "Stock Reconciliation",
    "Borrowing",
    "EPS",
)


def canonical_sheet_name(name: str | None) -> str:
    """Compare sheet titles ignoring incidental trailing spaces in the sample workbook."""
    return (name or "").strip()


def sheet_names_equal(left: str | None, right: str | None) -> bool:
    return canonical_sheet_name(left) == canonical_sheet_name(right)

METADATA_VERSION = 1

_PAGE_SETUP_ATTRS = (
    "orientation",
    "paperSize",
    "scale",
    "fitToPage",
    "fitToWidth",
    "fitToHeight",
    "firstPageNumber",
    "useFirstPageNumber",
    "pageOrder",
    "horizontalDpi",
    "verticalDpi",
    "copies",
    "draft",
    "blackAndWhite",
    "paperHeight",
    "paperWidth",
)

_MARGIN_ATTRS = ("left", "right", "top", "bottom", "header", "footer")


def file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _json_safe(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, bytes):
        return value.hex()
    return str(value)


def _color_to_dict(color: Any) -> dict[str, Any] | None:
    if color is None or color is False:
        return None
    payload: dict[str, Any] = {}
    for attr in ("type", "rgb", "theme", "indexed", "tint", "auto"):
        if not hasattr(color, attr):
            continue
        raw = getattr(color, attr)
        if raw is None:
            continue
        payload[attr] = _json_safe(raw)
    return payload or None


def _font_to_dict(font: Any) -> dict[str, Any] | None:
    if font is None:
        return None
    return {
        "name": font.name,
        "size": font.size,
        "bold": bool(font.bold),
        "italic": bool(font.italic),
        "underline": font.underline,
        "strike": bool(font.strike),
        "vertAlign": font.vertAlign,
        "color": _color_to_dict(font.color),
        "scheme": getattr(font, "scheme", None),
        "charset": getattr(font, "charset", None),
        "family": getattr(font, "family", None),
    }


def _alignment_to_dict(alignment: Any) -> dict[str, Any] | None:
    if alignment is None:
        return None
    return {
        "horizontal": alignment.horizontal,
        "vertical": alignment.vertical,
        "wrap_text": bool(alignment.wrap_text),
        "shrink_to_fit": bool(alignment.shrinkToFit),
        "indent": alignment.indent,
        "text_rotation": alignment.textRotation,
        "reading_order": getattr(alignment, "readingOrder", None),
    }


def _side_to_dict(side: Any) -> dict[str, Any] | None:
    if side is None or (not side.style and not side.color):
        return None
    return {"style": side.style, "color": _color_to_dict(side.color)}


def _border_to_dict(border: Any) -> dict[str, Any] | None:
    if border is None:
        return None
    payload = {
        "left": _side_to_dict(border.left),
        "right": _side_to_dict(border.right),
        "top": _side_to_dict(border.top),
        "bottom": _side_to_dict(border.bottom),
        "diagonal": _side_to_dict(border.diagonal),
        "diagonalUp": bool(getattr(border, "diagonalUp", False)),
        "diagonalDown": bool(getattr(border, "diagonalDown", False)),
        "outline": bool(getattr(border, "outline", True)),
    }
    if not any(payload[key] for key in ("left", "right", "top", "bottom", "diagonal")):
        return None
    return payload


def _fill_to_dict(fill: Any) -> dict[str, Any] | None:
    if fill is None or not getattr(fill, "fill_type", None):
        return None
    return {
        "fill_type": fill.fill_type,
        "patternType": getattr(fill, "patternType", None),
        "fgColor": _color_to_dict(getattr(fill, "fgColor", None)),
        "bgColor": _color_to_dict(getattr(fill, "bgColor", None)),
        "start_color": _color_to_dict(getattr(fill, "start_color", None)),
        "end_color": _color_to_dict(getattr(fill, "end_color", None)),
    }


def _protection_to_dict(protection: Any) -> dict[str, Any] | None:
    if protection is None:
        return None
    return {
        "locked": bool(protection.locked),
        "hidden": bool(protection.hidden),
    }


def _formula_from_cell(cell: Cell) -> str | None:
    value = cell.value
    if isinstance(value, ArrayFormula):
        text = value.text if isinstance(value.text, str) else str(value.text)
        return text
    if isinstance(value, DataTableFormula):
        return str(value)
    if isinstance(value, str) and value.startswith("="):
        return value
    if cell.data_type == "f" and value is not None:
        return str(value)
    return None


def _cell_is_blank(cell: Cell) -> bool:
    return cell.value is None and _formula_from_cell(cell) is None


def _has_non_default_style(cell: Cell) -> bool:
    font = cell.font
    if font:
        if font.bold or font.italic or font.strike or font.underline not in (None, "none"):
            return True
        if font.size not in (None, 11):
            return True
        if font.name not in (None, "Calibri"):
            return True
    if cell.number_format not in (None, "General"):
        return True
    if _border_to_dict(cell.border) or _fill_to_dict(cell.fill):
        return True
    alignment = cell.alignment
    if alignment and any(
        getattr(alignment, attr)
        for attr in ("horizontal", "wrap_text", "shrinkToFit", "indent", "textRotation")
    ):
        return True
    return False


def _serialize_cell(cell: Cell, *, in_merged_range: bool) -> dict[str, Any]:
    formula = _formula_from_cell(cell)
    blank = _cell_is_blank(cell)
    payload: dict[str, Any] = {
        "coordinate": cell.coordinate,
        "blank": blank,
        "data_type": cell.data_type,
        "value": None if formula else _json_safe(cell.value),
        "formula": formula,
        "number_format": cell.number_format,
        "font": _font_to_dict(cell.font),
        "font_size": cell.font.size if cell.font else None,
        "bold": bool(cell.font.bold) if cell.font else False,
        "italic": bool(cell.font.italic) if cell.font else False,
        "alignment": _alignment_to_dict(cell.alignment),
        "border": _border_to_dict(cell.border),
        "fill": _fill_to_dict(cell.fill),
        "protection": _protection_to_dict(cell.protection),
        "hyperlink": cell.hyperlink.target if cell.hyperlink else None,
        "comment": cell.comment.text if cell.comment else None,
        "in_merged_range": in_merged_range,
        "layout_placeholder": blank,
    }
    return payload


def _merged_coordinates(sheet: Worksheet) -> set[str]:
    coordinates: set[str] = set()
    for merged in sheet.merged_cells.ranges:
        for row in range(merged.min_row, merged.max_row + 1):
            for column in range(merged.min_col, merged.max_col + 1):
                coordinates.add(f"{get_column_letter(column)}{row}")
    return coordinates


def _row_dimensions(sheet: Worksheet) -> dict[str, dict[str, Any]]:
    rows: dict[str, dict[str, Any]] = {}
    for index, dim in sheet.row_dimensions.items():
        height = dim.height
        hidden = bool(dim.hidden)
        outline = dim.outlineLevel
        if height is None and not hidden and not outline:
            continue
        rows[str(index)] = {
            "height": height,
            "hidden": hidden,
            "outline_level": outline,
            "collapsed": bool(getattr(dim, "collapsed", False)),
        }
    return rows


def _column_dimensions(sheet: Worksheet) -> dict[str, dict[str, Any]]:
    columns: dict[str, dict[str, Any]] = {}
    for letter, dim in sheet.column_dimensions.items():
        width = dim.width
        hidden = bool(dim.hidden)
        if width is None and not hidden:
            continue
        columns[str(letter)] = {
            "width": width,
            "hidden": hidden,
            "outline_level": dim.outlineLevel,
            "best_fit": bool(getattr(dim, "bestFit", False)),
        }
    return columns


def _page_setup(sheet: Worksheet) -> dict[str, Any]:
    setup = sheet.page_setup
    payload = {attr: _json_safe(getattr(setup, attr, None)) for attr in _PAGE_SETUP_ATTRS}
    margins = sheet.page_margins
    payload["margins"] = {attr: getattr(margins, attr, None) for attr in _MARGIN_ATTRS}
    payload["print_titles"] = {
        "rows": sheet.print_title_rows,
        "columns": sheet.print_title_cols,
    }
    payload["print_options"] = {
        "horizontalCentered": getattr(sheet.print_options, "horizontalCentered", None),
        "verticalCentered": getattr(sheet.print_options, "verticalCentered", None),
        "headings": getattr(sheet.print_options, "headings", None),
        "gridLines": getattr(sheet.print_options, "gridLines", None),
    }
    payload["header"] = _header_footer_to_dict(getattr(sheet, "oddHeader", None))
    payload["footer"] = _header_footer_to_dict(getattr(sheet, "oddFooter", None))
    payload["even_header"] = _header_footer_to_dict(getattr(sheet, "evenHeader", None))
    payload["even_footer"] = _header_footer_to_dict(getattr(sheet, "evenFooter", None))
    payload["first_header"] = _header_footer_to_dict(getattr(sheet, "firstHeader", None))
    payload["first_footer"] = _header_footer_to_dict(getattr(sheet, "firstFooter", None))
    return payload


def _header_footer_to_dict(block: Any) -> dict[str, str | None]:
    if block is None:
        return {"left": None, "center": None, "right": None}

    def _part(name: str) -> str | None:
        part = getattr(block, name, None)
        if part is None:
            return None
        text = getattr(part, "text", None)
        return str(text) if text else None

    return {"left": _part("left"), "center": _part("center"), "right": _part("right")}


def _sheet_view(sheet: Worksheet) -> dict[str, Any]:
    view = sheet.sheet_view
    return {
        "show_grid_lines": getattr(view, "showGridLines", None),
        "show_row_col_headers": getattr(view, "showRowColHeaders", None),
        "zoom_scale": getattr(view, "zoomScale", None),
        "zoom_scale_normal": getattr(view, "zoomScaleNormal", None),
        "view": getattr(view, "view", None),
        "tab_selected": getattr(view, "tabSelected", None),
    }


def _inspect_sheet(sheet: Worksheet, index: int) -> dict[str, Any]:
    merged = [str(item) for item in sheet.merged_cells.ranges]
    merged_cells = _merged_coordinates(sheet)
    max_row = sheet.max_row or 1
    max_column = sheet.max_column or 1
    cells: dict[str, dict[str, Any]] = {}
    blank_layout_cells: list[str] = []
    formulas: dict[str, str] = {}
    row_heights = _row_dimensions(sheet)
    column_widths = _column_dimensions(sheet)

    existing_cells = getattr(sheet, "_cells", None)
    if isinstance(existing_cells, dict) and existing_cells:
        iterable = existing_cells.values()
    else:
        iterable = (
            cell
            for row in sheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_column)
            for cell in row
        )
    for cell in iterable:
        if not isinstance(cell, Cell):
            continue
        in_merged = cell.coordinate in merged_cells
        blank = _cell_is_blank(cell)
        if blank and not in_merged and not cell.has_style:
            continue
        relevant = not blank or in_merged or _has_non_default_style(cell)
        if not relevant:
            continue
        cells[cell.coordinate] = _serialize_cell(cell, in_merged_range=in_merged)
        if blank:
            blank_layout_cells.append(cell.coordinate)
        formula = _formula_from_cell(cell)
        if formula:
            formulas[cell.coordinate] = formula

    return {
        "name": sheet.title,
        "index": index,
        "order": index,
        "sheet_state": sheet.sheet_state,
        "hidden": sheet.sheet_state != "visible",
        "dimensions": sheet.dimensions,
        "max_row": max_row,
        "max_column": max_column,
        "freeze_panes": str(sheet.freeze_panes) if sheet.freeze_panes else None,
        "print_area": sheet.print_area or None,
        "merged_cells": merged,
        "row_heights": row_heights,
        "column_widths": column_widths,
        "hidden_rows": [int(key) for key, payload in row_heights.items() if payload.get("hidden")],
        "hidden_columns": [letter for letter, payload in column_widths.items() if payload.get("hidden")],
        "page_setup": _page_setup(sheet),
        "sheet_view": _sheet_view(sheet),
        "auto_filter": sheet.auto_filter.ref if sheet.auto_filter and sheet.auto_filter.ref else None,
        "formulas": formulas,
        "blank_layout_cells": blank_layout_cells,
        "cells": cells,
    }


def _defined_names(workbook: Workbook) -> list[dict[str, Any]]:
    names: list[dict[str, Any]] = []
    defined = workbook.defined_names
    for name in defined.values():
        if not isinstance(name, DefinedName):
            continue
        names.append(
            {
                "name": name.name,
                "attr_text": name.attr_text,
                "value": getattr(name, "value", None),
                "hidden": bool(name.hidden),
                "comment": name.comment,
            }
        )
    return names


@dataclass(slots=True)
class TemplateMetadata:
    """Machine-readable snapshot of the Excel template layout."""

    version: int
    source_path: str
    source_sha256: str
    sheet_names: list[str]
    sheet_order: list[str]
    expected_sheet_names: list[str]
    sheets: list[dict[str, Any]]
    defined_names: list[dict[str, Any]]
    named_styles: list[str]
    extra: dict[str, Any] = field(default_factory=dict)

    def to_dict(self) -> dict[str, Any]:
        return {
            "version": self.version,
            "source_path": self.source_path,
            "source_sha256": self.source_sha256,
            "sheet_names": self.sheet_names,
            "sheet_order": self.sheet_order,
            "expected_sheet_names": self.expected_sheet_names,
            "defined_names": self.defined_names,
            "named_styles": self.named_styles,
            "workbook_formulas": self.defined_names,
            "sheets": self.sheets,
            **self.extra,
        }


class TemplateManager:
    """Load, inspect, and copy the Division I Schedule III workbook template."""

    def __init__(self, template_path: Path | None = None) -> None:
        self.template_path = Path(template_path) if template_path else settings.template_path
        if not self.template_path.exists():
            raise FileNotFoundError(f"Excel template not found: {self.template_path}")
        if self.template_path.suffix.lower() not in {".xlsx", ".xlsm"}:
            raise ValueError(f"Unsupported template type: {self.template_path}")

    def load(self, *, data_only: bool = False) -> Workbook:
        """Open the original template as read-only in the application sense.

        The returned workbook is never saved back to ``template_path``.
        """
        return load_workbook(
            self.template_path,
            data_only=data_only,
            keep_vba=self.template_path.suffix.lower() == ".xlsm",
            rich_text=True,
        )

    def inspect(self) -> TemplateMetadata:
        """Inspect layout, formatting, formulas, and values without changing the file."""
        before = file_sha256(self.template_path)
        workbook = self.load(data_only=False)
        try:
            sheets = [_inspect_sheet(sheet, index) for index, sheet in enumerate(workbook.worksheets)]
            metadata = TemplateMetadata(
                version=METADATA_VERSION,
                source_path=str(self.template_path),
                source_sha256=before,
                sheet_names=list(workbook.sheetnames),
                sheet_order=list(workbook.sheetnames),
                expected_sheet_names=list(EXPECTED_SHEET_NAMES),
                sheets=sheets,
                defined_names=_defined_names(workbook),
                named_styles=list(getattr(workbook, "style_names", [])),
                extra={
                    "active_sheet": workbook.active.title if workbook.active is not None else None,
                    "excel_base_date": str(getattr(workbook, "epoch", "")),
                },
            )
        finally:
            workbook.close()

        after = file_sha256(self.template_path)
        if after != before:
            raise RuntimeError("Template inspection modified the original workbook.")
        return metadata

    def export_metadata(self, output_path: Path | None = None) -> Path:
        """Write template metadata JSON. Does not modify the original workbook."""
        destination = Path(output_path) if output_path else settings.output_dir / "template_metadata.json"
        destination.parent.mkdir(parents=True, exist_ok=True)
        payload = self.inspect().to_dict()
        destination.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
        return destination

    def copy_template(self, destination: Path | None = None) -> Path:
        """Copy the original template so later writes never touch the source file."""
        target = (
            Path(destination)
            if destination
            else settings.output_dir / self.template_path.name
        )
        if target.resolve() == self.template_path.resolve():
            raise ValueError("Refusing to copy the template onto itself.")
        target.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(self.template_path, target)
        return target

    def missing_expected_sheets(self) -> list[str]:
        workbook = self.load()
        try:
            present = {canonical_sheet_name(name) for name in workbook.sheetnames}
        finally:
            workbook.close()
        return [name for name in EXPECTED_SHEET_NAMES if canonical_sheet_name(name) not in present]


def main() -> int:
    manager = TemplateManager()
    metadata_path = manager.export_metadata()
    metadata = json.loads(metadata_path.read_text(encoding="utf-8"))
    print(f"Template: {manager.template_path}")
    print(f"SHA-256: {metadata['source_sha256']}")
    print("Sheet order:")
    for index, name in enumerate(metadata["sheet_order"]):
        print(f"  {index + 1}. {name}")
    missing = manager.missing_expected_sheets()
    if missing:
        print("Expected sheets not present in this workbook:")
        for name in missing:
            print(f"  - {name}")
    print(f"Metadata JSON: {metadata_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

