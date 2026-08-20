"""Replace sample-company branding in the ICAI template with uploaded company details."""

from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Iterable

from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.config import settings


def _resolve_sheet_name(workbook: Workbook, sheet_name: str) -> str | None:
    if sheet_name in workbook.sheetnames:
        return sheet_name
    wanted = sheet_name.strip()
    for name in workbook.sheetnames:
        if name.strip() == wanted:
            return name
    return None


def normalize_company_name(value: str | None) -> str:
    if not value:
        return ""
    cleaned = re.sub(r"[^a-z0-9]+", " ", value.lower())
    return " ".join(cleaned.split())


def template_company_names(config_path: Path | None = None) -> tuple[str, ...]:
    path = Path(config_path) if config_path else settings.mapping_config_path
    if not path.exists():
        return ("ABC INDIA LIMITED", "ABC India Limited", "ABC & Company")
    payload = json.loads(path.read_text(encoding="utf-8"))
    names: list[str] = []
    primary = payload.get("template_company_name")
    if isinstance(primary, str) and primary.strip():
        names.append(primary.strip())
    for item in payload.get("template_company_aliases") or []:
        if isinstance(item, str) and item.strip():
            names.append(item.strip())
    for item in payload.get("template_sample_phrases") or []:
        if isinstance(item, str) and item.strip():
            names.append(item.strip())
    if not names:
        names.extend(["ABC INDIA LIMITED", "ABC India Limited", "ABC & Company"])
    deduped: list[str] = []
    seen: set[str] = set()
    for name in names:
        key = normalize_company_name(name)
        if key and key not in seen:
            seen.add(key)
            deduped.append(name)
    return tuple(deduped)


def company_name_from_placements(placements: Iterable[dict]) -> str | None:
    for placement in placements:
        if placement.get("field_key") != "company_name":
            continue
        value = placement.get("extracted_value")
        if isinstance(value, str) and value.strip():
            return value.strip()
    return None


def apply_company_branding(
    workbook: Workbook,
    company_name: str,
    *,
    template_names: Iterable[str] | None = None,
    full_workbook: bool = False,
) -> list[dict[str, str]]:
    """Replace the ICAI sample company name throughout the workbook."""
    if not company_name.strip():
        return []
    templates = list(template_names or template_company_names())
    template_norms = {normalize_company_name(name) for name in templates}
    written: list[dict[str, str]] = []
    for sheet in workbook.worksheets:
        if full_workbook:
            written.extend(_replace_on_sheet(sheet, templates, template_norms, company_name))
        else:
            for row in range(1, 11):
                for col in range(1, 6):
                    cell = sheet.cell(row=row, column=col)
                    value = cell.value
                    if not isinstance(value, str) or not value.strip():
                        continue
                    updated = _replace_company_text(value, templates, template_norms, company_name)
                    if updated != value:
                        cell.value = updated
                        written.append({"sheet": sheet.title, "cell": cell.coordinate, "value": updated})
    return written


def _replace_on_sheet(
    sheet: Worksheet,
    templates: list[str],
    template_norms: set[str],
    company_name: str,
) -> list[dict[str, str]]:
    written: list[dict[str, str]] = []
    for row in sheet.iter_rows():
        for cell in row:
            value = cell.value
            if not isinstance(value, str) or not value.strip():
                continue
            updated = _replace_company_text(value, templates, template_norms, company_name)
            if updated != value:
                cell.value = updated
                written.append({"sheet": sheet.title, "cell": cell.coordinate, "value": updated})
    return written


def write_company_header_placements(
    workbook: Workbook,
    company_name: str,
    *,
    template_names: Iterable[str] | None = None,
) -> list[dict[str, str]]:
    """Ensure the primary face-sheet header shows the uploaded company name."""
    templates = list(template_names or template_company_names())
    template_norms = {normalize_company_name(name) for name in templates}
    sheet_name = _resolve_sheet_name(workbook, "BS PnL")
    if sheet_name is None:
        return []
    sheet: Worksheet = workbook[sheet_name]
    target = sheet["A2"]
    current = target.value
    if isinstance(current, str):
        updated = _replace_company_text(current, templates, template_norms, company_name)
    else:
        updated = company_name
    if updated != current:
        target.value = updated
        return [{"sheet": sheet_name, "cell": "A2", "value": updated}]
    return []


def _replace_company_text(
    value: str,
    templates: list[str],
    template_norms: set[str],
    company_name: str,
) -> str:
    stripped = value.strip()
    if normalize_company_name(stripped) in template_norms:
        return company_name
    updated = value
    for template in sorted(templates, key=len, reverse=True):
        pattern = re.compile(re.escape(template), flags=re.IGNORECASE)
        if pattern.search(updated):
            updated = pattern.sub(company_name, updated)
    if re.search(r"\bFor\s+ABC\s*&\s*Company\b", updated, flags=re.IGNORECASE):
        updated = re.sub(r"\bFor\s+ABC\s*&\s*Company\b", "For ...........................", updated, flags=re.IGNORECASE)
    return updated
