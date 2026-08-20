"""Generate Schedule III workbooks by filling a copy of the sample template.

The original file ``templates/Financial Statements_Sample.xlsx`` is never
opened for save. Every run copies it, writes into the copy, then validates
that structure and formulas were preserved.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.config import settings
from app.excel.company_branding import (
    apply_company_branding,
    company_name_from_placements,
    template_company_names,
)
from app.excel.formatting import write_formula, write_text, write_value
from app.excel.template_manager import TemplateManager, file_sha256
from app.excel.trial_balance_workbook import remove_excluded_sheets, sanitize_trial_balance_workbook

DEFAULT_OUTPUT_FILENAME = "Financial_Statements_Generated.xlsx"


@dataclass(slots=True)
class WorkbookValidation:
    """Structural comparison of a generated workbook against the master template."""

    ok: bool
    errors: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    original_hash: str | None = None
    generated_path: str | None = None

    def to_dict(self) -> dict[str, Any]:
        return {
            "ok": self.ok,
            "errors": list(self.errors),
            "warnings": list(self.warnings),
            "original_hash": self.original_hash,
            "generated_path": self.generated_path,
        }


@dataclass(slots=True)
class GenerationResult:
    path: Path
    validation: WorkbookValidation
    written: list[dict[str, Any]] = field(default_factory=list)
    skipped: list[dict[str, Any]] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        return {
            "path": str(self.path),
            "validation": self.validation.to_dict(),
            "written": list(self.written),
            "skipped": list(self.skipped),
        }


def copy_template(
    destination: str | Path | None = None,
    *,
    template_path: str | Path | None = None,
) -> Path:
    """Copy the master template. Refuses to write onto the original file."""
    manager = TemplateManager(Path(template_path) if template_path else None)
    target = Path(destination) if destination else settings.output_dir / DEFAULT_OUTPUT_FILENAME
    return manager.copy_template(target)


def validate_workbook(
    generated_path: str | Path,
    template_path: str | Path | None = None,
    *,
    allowed_formula_overwrites: Iterable[tuple[str, str]] | None = None,
    allowed_value_cells: Iterable[tuple[str, str]] | None = None,
    expected_original_hash: str | None = None,
    allowed_missing_sheets: Iterable[str] | None = None,
) -> WorkbookValidation:
    """Confirm the generated file still matches the template's layout and formatting."""
    from app.excel.template_comparator import compare_workbooks

    generated = Path(generated_path)
    manager = TemplateManager(Path(template_path) if template_path else None)
    original_hash = file_sha256(manager.template_path)
    errors: list[str] = []
    warnings: list[str] = []

    if expected_original_hash and original_hash != expected_original_hash:
        errors.append("The original template was modified during generation.")

    report = compare_workbooks(
        manager.template_path,
        generated,
        allowed_formula_overwrites=allowed_formula_overwrites,
        allowed_value_cells=allowed_value_cells,
        allowed_missing_sheets=allowed_missing_sheets,
    )
    errors.extend(item.message for item in report.errors)
    warnings.extend(item.message for item in report.warnings)

    return WorkbookValidation(
        ok=not errors,
        errors=errors,
        warnings=warnings,
        original_hash=original_hash,
        generated_path=str(generated),
    )


class WorkbookGenerator:
    """Copy the ICAI sample workbook and populate the copy only."""

    def __init__(
        self,
        output_dir: Path | None = None,
        *,
        template_path: Path | None = None,
    ) -> None:
        self.output_dir = Path(output_dir) if output_dir else settings.output_dir
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.template_manager = TemplateManager(template_path)
        self.last_result: GenerationResult | None = None

    @property
    def template_path(self) -> Path:
        return self.template_manager.template_path

    def copy_template(self, destination: str | Path | None = None) -> Path:
        target = Path(destination) if destination else self.output_dir / DEFAULT_OUTPUT_FILENAME
        return copy_template(target, template_path=self.template_path)

    def generate(
        self,
        mapped_data: dict | None = None,
        output_filename: str | None = None,
    ) -> Path:
        """Create ``output/Financial_Statements_Generated.xlsx`` from mapped placements."""
        return self.generate_detailed(mapped_data or {}, output_filename=output_filename).path

    def generate_detailed(
        self,
        mapped_data: dict | None = None,
        output_filename: str | None = None,
    ) -> GenerationResult:
        original_hash = file_sha256(self.template_path)
        destination = self._resolve_output_path(output_filename)
        copied = self.copy_template(destination)
        if file_sha256(self.template_path) != original_hash:
            raise RuntimeError("Copying the template modified the original workbook.")

        placements = _placements_from(mapped_data or {})
        generation_mode = str((mapped_data or {}).get("generation_mode") or "")
        exclude_sheets = list((mapped_data or {}).get("exclude_sheets") or [])
        company_name = company_name_from_placements(placements) or (mapped_data or {}).get("company_name")
        period_label = (mapped_data or {}).get("reporting_period")
        workbook = load_workbook(
            copied,
            data_only=False,
            keep_vba=copied.suffix.lower() == ".xlsm",
            rich_text=True,
        )
        written: list[dict[str, Any]] = []
        skipped: list[dict[str, Any]] = []
        overwritten_formulas: list[tuple[str, str]] = []
        try:
            if generation_mode == "trial_balance":
                removed = remove_excluded_sheets(workbook, exclude_sheets or ("Cash Flow",))
                for sheet_name in removed:
                    written.append(
                        {
                            "sheet": sheet_name,
                            "cell": None,
                            "value": None,
                            "action": "write",
                            "written": True,
                            "reason": "removed_sheet",
                            "field_key": None,
                            "period": None,
                        }
                    )
            if company_name:
                templates = template_company_names()
                for record in apply_company_branding(
                    workbook,
                    str(company_name),
                    template_names=templates,
                    full_workbook=generation_mode == "trial_balance",
                ):
                    written.append(
                        {
                            "sheet": record["sheet"],
                            "cell": record["cell"],
                            "value": record["value"],
                            "action": "write",
                            "written": True,
                            "reason": "company_branding",
                            "overwrite_formula": False,
                            "field_key": "company_name",
                            "period": "current",
                        }
                    )
                if generation_mode == "trial_balance":
                    for record in sanitize_trial_balance_workbook(
                        workbook,
                        company_name=str(company_name),
                        period_label=str(period_label) if period_label else None,
                        template_names=templates,
                    ):
                        written.append(
                            {
                                "sheet": record["sheet"],
                                "cell": record["cell"],
                                "value": record["value"],
                                "action": "write",
                                "written": True,
                                "reason": "trial_balance_sanitize",
                                "overwrite_formula": False,
                                "field_key": "company_name",
                                "period": "current",
                            }
                        )
            for placement in placements:
                record = self._apply_placement(workbook, placement)
                if record.get("written"):
                    written.append(record)
                    if record.get("overwrite_formula"):
                        overwritten_formulas.append((record["sheet"], record["cell"]))
                else:
                    skipped.append(record)
            workbook.save(copied)
        finally:
            workbook.close()

        if file_sha256(self.template_path) != original_hash:
            raise RuntimeError("Saving the generated workbook modified the original template.")

        validation = validate_workbook(
            copied,
            self.template_path,
            allowed_formula_overwrites=overwritten_formulas,
            allowed_value_cells=[
                (record["sheet"], record["cell"])
                for record in written
                if record.get("sheet") and record.get("cell")
            ],
            expected_original_hash=original_hash,
            allowed_missing_sheets=exclude_sheets if generation_mode == "trial_balance" else None,
        )
        result = GenerationResult(
            path=copied,
            validation=validation,
            written=written,
            skipped=skipped,
        )
        self.last_result = result
        return result

    def validate(self, generated_path: str | Path) -> WorkbookValidation:
        return validate_workbook(generated_path, self.template_path)

    def _resolve_output_path(self, output_filename: str | None) -> Path:
        name = output_filename or DEFAULT_OUTPUT_FILENAME
        path = Path(name)
        if path.is_absolute() or path.parent != Path("."):
            destination = path if path.is_absolute() else self.output_dir / path
        else:
            destination = self.output_dir / path.name
        if destination.resolve() == self.template_path.resolve():
            raise ValueError("Refusing to generate onto the original template.")
        destination.parent.mkdir(parents=True, exist_ok=True)
        return destination

    def _apply_placement(self, workbook: Workbook, placement: dict[str, Any]) -> dict[str, Any]:
        sheet_name = placement.get("excel_sheet")
        cell = placement.get("excel_cell")
        action = placement.get("action") or "write"
        value = placement.get("extracted_value")
        overwrite = bool(placement.get("overwrite_formula"))
        record = {
            "sheet": sheet_name,
            "cell": cell,
            "value": value,
            "action": action,
            "written": False,
            "reason": None,
            "overwrite_formula": overwrite,
            "field_key": placement.get("field_key"),
            "period": placement.get("period"),
        }
        if action != "write":
            record["reason"] = action
            return record
        if placement.get("review_status") not in (None, "approved"):
            record["reason"] = "not_approved"
            record["action"] = "not_approved"
            return record
        if not sheet_name or not cell:
            record["reason"] = "missing_destination"
            return record
        resolved_sheet = _resolve_sheet_name(workbook, sheet_name)
        if resolved_sheet is None:
            record["reason"] = "missing_sheet"
            return record
        record["sheet"] = resolved_sheet
        worksheet: Worksheet = workbook[resolved_sheet]
        kind = placement.get("value_role") or _infer_write_kind(value)
        if kind == "formula":
            formula = placement.get("formula") or value
            result = write_formula(worksheet, str(cell), str(formula), overwrite_formula=overwrite)
        elif kind == "text":
            result = write_text(worksheet, str(cell), None if value is None else str(value), overwrite_formula=overwrite)
        else:
            numeric = value if isinstance(value, (int, float)) else _as_number(value)
            if numeric is None and value is not None:
                result = write_text(worksheet, str(cell), str(value), overwrite_formula=overwrite)
            else:
                result = write_value(worksheet, str(cell), numeric, overwrite_formula=overwrite)
        record["cell"] = result.coordinate
        record["written"] = result.written
        record["reason"] = result.reason
        return record


def _resolve_sheet_name(workbook: Workbook, sheet_name: str) -> str | None:
    """Match the sample's sheet titles, including trailing spaces such as ``Cash Flow ``."""
    if sheet_name in workbook.sheetnames:
        return sheet_name
    wanted = sheet_name.strip()
    for name in workbook.sheetnames:
        if name.strip() == wanted:
            return name
    return None


def _placements_from(mapped_data: dict[str, Any]) -> list[dict[str, Any]]:
    if not mapped_data:
        return []
    placements = mapped_data.get("placements")
    if isinstance(placements, list):
        return [item for item in placements if isinstance(item, dict)]
    return []


def _infer_write_kind(value: Any) -> str:
    if isinstance(value, str) and value.startswith("="):
        return "formula"
    if isinstance(value, str):
        return "text"
    return "value"


def _as_number(value: Any) -> float | int | None:
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, (int, float)):
        return value
    return None
