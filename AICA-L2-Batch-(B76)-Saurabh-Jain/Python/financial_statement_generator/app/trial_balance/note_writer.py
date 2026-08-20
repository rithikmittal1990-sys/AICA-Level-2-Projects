"""
Note Writer — renders NoteDataModel into Excel Notes sheets.

Strategy:
 1. For every Notes sheet: wipe ALL numeric sample data and ALL placeholder text.
 2. Write only values from the NoteDataModel (source = TRIAL_BALANCE or CALCULATED).
 3. Preserve structural row labels (headings, sub-headings).
 4. Write N.A. for comparative columns everywhere.
 5. Write NA_TEXT for any disclosure that cannot be sourced from TB.
"""

from __future__ import annotations

import re
from typing import TYPE_CHECKING

from openpyxl.styles import Font

from app.trial_balance.note_data_model import NA_TEXT, MGMT_TEXT, NoteDataModel, NoteEntry

if TYPE_CHECKING:
    from openpyxl import Workbook
    from openpyxl.worksheet.worksheet import Worksheet

_SAMPLE_RE = re.compile(
    r"\b(XYZ Limited(?! Private)|XYZ Private Limited|ABCDE|Punit Kumar Sarda"
    r"|Object Clause|99,75,360|9,975,360|8,312,800|1,662,560|1,496,304"
    r"|30991)\b",
    re.IGNORECASE,
)


# --------------------------------------------------------------------------- #
# Helpers
# --------------------------------------------------------------------------- #

def _ws(wb: "Workbook", name: str) -> "Worksheet | None":
    for n in wb.sheetnames:
        if n.strip() == name.strip():
            return wb[n]
    return None


def _set(ws: "Worksheet", coord: str, value, changes: list[str]) -> None:
    from openpyxl.cell.cell import MergedCell
    cell = ws[coord]
    if isinstance(cell, MergedCell):
        return
    if cell.value != value:
        changes.append(f"{ws.title}!{coord}: {repr(str(cell.value)[:30])} → {repr(str(value)[:40])}")
        cell.value = value


def _wipe_numerics(ws: "Worksheet", col_letters: list[str], rows: range) -> None:
    """Clear all numeric values in given columns and rows."""
    for row in rows:
        for col in col_letters:
            from openpyxl.cell.cell import MergedCell
            cell = ws[f"{col}{row}"]
            if isinstance(cell, MergedCell):
                continue
            if isinstance(cell.value, (int, float)):
                cell.value = None


def _wipe_sample_text(ws: "Worksheet", changes: list[str]) -> None:
    """Replace sample entity names and placeholder text across the sheet."""
    from openpyxl.cell.cell import MergedCell
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            if isinstance(cell.value, str) and _SAMPLE_RE.search(cell.value):
                old = cell.value
                cell.value = NA_TEXT
                changes.append(f"{ws.title}!{cell.coordinate}: sample text cleared")


def _na_col(ws: "Worksheet", col: str, rows: list[int]) -> None:
    """Set prior-year column to N.A."""
    from openpyxl.cell.cell import MergedCell
    for row in rows:
        cell = ws[f"{col}{row}"]
        if isinstance(cell, MergedCell):
            continue
        if cell.value != "N.A.":
            cell.value = "N.A."


# --------------------------------------------------------------------------- #
# Master entry point
# --------------------------------------------------------------------------- #

def write_notes_from_model(workbook: "Workbook", model: NoteDataModel) -> list[str]:
    changes: list[str] = []
    _write_note1_2(workbook, model, changes)
    _write_note3(workbook, model, changes)
    _write_note4_12(workbook, model, changes)
    _write_note12_ppe(workbook, model, changes)
    _write_cwip(workbook, model, changes)
    _write_note13_20(workbook, model, changes)
    _write_note20_31(workbook, model, changes)
    _write_ebp(workbook, model, changes)
    _write_other_notes(workbook, model, changes)
    _write_stock_recon(workbook, model, changes)
    _write_borrowing(workbook, model, changes)
    _write_eps(workbook, model, changes)
    return changes


# --------------------------------------------------------------------------- #
# Note 1-2
# --------------------------------------------------------------------------- #

def _write_note1_2(wb: "Workbook", m: NoteDataModel, ch: list[str]) -> None:
    ws = _ws(wb, "Note 1-2")
    if not ws:
        return
    _set(ws, "A3",
         f"Notes forming part of the Financial Statements for the year ended {m.reporting_date}", ch)
    _set(ws, "B10",
         f"{m.company_name} ('the Company') — financial statements prepared solely from the "
         f"uploaded trial balance for {m.period_label}.", ch)
    if ws["A77"].value == "2.1":
        _set(ws, "A77", "2.9", ch)
    _wipe_sample_text(ws, ch)


# --------------------------------------------------------------------------- #
# Note 3 — Share Capital
# --------------------------------------------------------------------------- #

def _write_note3(wb: "Workbook", m: NoteDataModel, ch: list[str]) -> None:
    ws = _ws(wb, "Note 3 (Share Capital)")
    if not ws:
        return
    note = m.by_note("3")

    # Dynamic column headers
    for coord, val in [
        ("F7",  f"As at {m.reporting_date}"), ("J7",  f"As at {m.prior_year_date} (N.A.)"),
        ("F18", f"As at {m.reporting_date}"), ("J18", f"As at {m.prior_year_date} (N.A.)"),
        ("F48", f"As at {m.reporting_date}"), ("K48", f"As at {m.prior_year_date} (N.A.)"),
        ("F63", f"As at {m.reporting_date}"), ("K63", f"As at {m.prior_year_date} (N.A.)"),
    ]:
        _set(ws, coord, val, ch)

    # Clear old data
    for coord in ("F12", "H12", "H13", "F20", "F21", "F22", "F23"):
        from openpyxl.cell.cell import MergedCell
        cell = ws[coord]
        if not isinstance(cell, MergedCell) and cell.value is not None:
            cell.value = None
            ch.append(f"Note 3!{coord}: cleared")

    if note:
        sc_items = [i for i in note.line_items if isinstance(i.amount, (int, float)) and not i.is_subtotal]
        sc_total  = note.total or 0.0
        n_shares  = int(sc_total / 10) if sc_total else None

        _set(ws, "B10", "Authorised share capital — Information not available from Trial Balance", ch)
        _set(ws, "B12",
             f"Equity shares of ₹10 each — issued, subscribed and fully paid up"
             + (f" ({n_shares:,} shares)" if n_shares else ""),
             ch)
        if n_shares:
            _set(ws, "F12", n_shares, ch)
            _set(ws, "H12", sc_total, ch)
        _set(ws, "H13", sc_total, ch)
        for coord in ("F20", "F21", "F22", "F23"):
            _set(ws, coord, NA_TEXT, ch)

    # Five-year history
    _set(ws, "F37", m.reporting_year, ch)
    for coord in ("H37", "J37", "L37", "N37"):
        _set(ws, coord, NA_TEXT, ch)
    for row in (39, 41, 43):
        for col in ("F", "H", "J", "L", "N"):
            from openpyxl.cell.cell import MergedCell
            cell = ws[f"{col}{row}"]
            if not isinstance(cell, MergedCell) and str(cell.value or "").strip() in {"Nil", "0", ""}:
                cell.value = NA_TEXT

    # Shareholder / promoter rows — wipe sample names
    for row in range(50, 75):
        for col in ("B",):
            from openpyxl.cell.cell import MergedCell
            cell = ws[f"{col}{row}"]
            if not isinstance(cell, MergedCell) and isinstance(cell.value, str):
                if _SAMPLE_RE.search(cell.value) or "XYZ" in cell.value:
                    cell.value = NA_TEXT
                    ch.append(f"Note 3!{col}{row}: sample name cleared")

    # Placeholder zeros in comparative
    for coord in ("O51", "O53", "M55", "O55", "O57", "O66", "O68", "O70", "O72"):
        from openpyxl.cell.cell import MergedCell
        cell = ws[coord]
        if not isinstance(cell, MergedCell) and str(cell.value or "").strip() in {"0", "Nil"}:
            cell.value = NA_TEXT

    _na_col(ws, "H", [8, 12, 13, 19, 23])
    _wipe_sample_text(ws, ch)


# --------------------------------------------------------------------------- #
# NOTE (4-12) — Notes 4 through 11
# --------------------------------------------------------------------------- #

def _write_note4_12(wb: "Workbook", m: NoteDataModel, ch: list[str]) -> None:
    ws = _ws(wb, "NOTE (4-12)")
    if not ws:
        return

    _set(ws, "A3",
         f"Notes to financial statements for the year ended at {m.reporting_date}", ch)

    # ---- Note 4: Reserves & Surplus -------------------------------- #
    res = m.by_note("4")
    if res:
        opening = next((i for i in res.line_items if "opening" in i.label.lower()), None)
        profit  = next((i for i in res.line_items if "profit" in i.label.lower()), None)
        closing = next((i for i in res.line_items if "closing" in i.label.lower()), None)
        _set(ws, "E10", opening.amount if opening and isinstance(opening.amount, (int, float)) else NA_TEXT, ch)
        _set(ws, "F10", "N.A.", ch)
        _set(ws, "E11", profit.amount if profit and isinstance(profit.amount, (int, float)) else NA_TEXT, ch)
        _set(ws, "F11", "N.A.", ch)
        _set(ws, "F12", "N.A.", ch)
        _set(ws, "F13", "N.A.", ch)
        _set(ws, "F14", "N.A.", ch)
        _set(ws, "F15", "N.A.", ch)
        _set(ws, "E16", opening.amount if opening and isinstance(opening.amount, (int, float)) else NA_TEXT, ch)
        _set(ws, "F16", "N.A.", ch)

    # ---- Note 5: Borrowings ---------------------------------------- #
    _set(ws, "B31", "Refer borrowing schedule", ch)
    _set(ws, "B72", "Refer borrowing schedule", ch)

    # ---- Note 6: Deferred Tax -------------------------------------- #
    _set(ws, "F35", f"As at {m.reporting_date}", ch)
    _set(ws, "F40", f"As at {m.prior_year_date} (N.A.)", ch)
    _set(ws, "D35", "N.A.", ch)
    _set(ws, "D40", "N.A.", ch)
    _set(ws, "E35", f"Recognised in P&L — FY {m.reporting_year}", ch)
    _set(ws, "E40", "N.A.", ch)

    # Wipe deferred tax numeric data
    for row in range(36, 44):
        _wipe_numerics(ws, ["C", "D", "E", "F", "G"], range(row, row + 1))

    dt = m.by_note("6")
    if dt:
        row = 36
        for item in dt.line_items:
            if isinstance(item.amount, (int, float)):
                _set(ws, f"F{row}", item.amount, ch)
                row += 1
            else:
                _set(ws, f"E{row}", item.amount, ch)
                row += 1

    # ---- Note 9: Trade Payables (Sundry Creditors ONLY) ------------ #
    tp = m.by_note("9")
    _set(ws, "G76", f"As at {m.reporting_date}", ch)
    _set(ws, "G87", f"As at {m.prior_year_date} (N.A.)", ch)
    # Wipe old values
    _wipe_numerics(ws, ["C", "D", "E", "F", "G", "H"], range(79, 97))
    if tp:
        # Write individual TP accounts starting at row 81 (Others)
        row = 81
        for item in tp.line_items:
            if isinstance(item.amount, (int, float)) and row <= 85:
                _set(ws, f"D{row}", item.amount, ch)
                _set(ws, f"B{row}", item.label, ch)
                row += 1
        _set(ws, "D85", tp.total if tp.total else NA_TEXT, ch)
    _set(ws, "B97", "* Related party trade payable details: "
         "Information is not determinable from the Trial Balance and requires additional management information.", ch)

    # ---- Note 10: Other Current Liabilities ------------------------ #
    ocl = m.by_note("10")
    _set(ws, "B101", "Particulars", ch)
    if ocl:
        data_rows = [(i.label, i.amount) for i in ocl.line_items
                     if isinstance(i.amount, (int, float))]
        for idx, (lbl, val) in enumerate(data_rows):
            r = 102 + idx
            _set(ws, f"B{r}", lbl, ch)
            _set(ws, f"E{r}", val, ch)
        _set(ws, f"E{102 + len(data_rows)}", ocl.total, ch)

    # ---- Note 11: Short-Term Provisions ---------------------------- #
    prov = m.by_note("11")
    if prov:
        data_rows = [(i.label, i.amount) for i in prov.line_items
                     if isinstance(i.amount, (int, float))]
        for idx, (lbl, val) in enumerate(data_rows):
            r = 117 + idx
            _set(ws, f"B{r}", lbl, ch)
            _set(ws, f"E{r}", val, ch)
        _set(ws, f"E{117 + len(data_rows)}", prov.total, ch)

    _wipe_sample_text(ws, ch)


# --------------------------------------------------------------------------- #
# Note 12 — PPE
# --------------------------------------------------------------------------- #

def _write_note12_ppe(wb: "Workbook", m: NoteDataModel, ch: list[str]) -> None:
    ws = _ws(wb, "Note 12 (PPE)")
    if not ws:
        return
    ppe = m.by_note("12")
    yr  = m.reporting_date[-4:]

    _set(ws, "B8", f"ON 01.04.{int(yr)-1}", ch)
    _set(ws, "E8", f"31.03.{yr}", ch)
    _set(ws, "F8", f"31.03.{int(yr)-1}", ch)
    _set(ws, "J8", f"ON 31.03.{yr}", ch)
    _set(ws, "K8", f"ON 31.03.{int(yr)-1}", ch)

    # Wipe ALL numeric data cells in PPE block; we re-write from model
    _wipe_numerics(ws, list("BCDEFGHIJK"), range(10, 40))

    if ppe:
        ppe_items = [i for i in ppe.line_items
                     if isinstance(i.amount, (int, float)) and not i.is_subtotal
                     and "tangible" not in i.label.lower() and "intangible" not in i.label.lower()
                     and "B." not in i.label and "A." not in i.label]
        ia_items  = [i for i in ppe.line_items
                     if isinstance(i.amount, (int, float)) and not i.is_subtotal
                     and "intangible" in i.label.lower()]
        ppe_items_clean = [i for i in ppe_items if "intangible" not in i.label.lower()]

        ppe_total_item = next((i for i in ppe.line_items if i.is_subtotal and "tangible" in i.label.lower()), None)
        ia_total_item  = next((i for i in ppe.line_items if i.is_subtotal and "intangible" in i.label.lower()), None)

        # Write individual asset rows starting at A12
        row = 12
        for item in ppe_items_clean:
            _set(ws, f"A{row}", item.label.strip(), ch)
            _set(ws, f"J{row}", item.amount, ch)  # NBV current year
            _set(ws, f"K{row}", "N.A.", ch)        # NBV prior year
            # Movement columns — not available
            for col in ("B", "C", "D", "E", "F", "G", "H", "I"):
                _set(ws, f"{col}{row}", NA_TEXT, ch)
            row += 1
        # Total A row
        _set(ws, "J20", ppe.total or (ppe_total_item.amount if ppe_total_item else NA_TEXT), ch)
        _set(ws, "K20", "N.A.", ch)

        # Intangible rows
        row = 25
        for item in ia_items:
            _set(ws, f"A{row}", item.label.strip(), ch)
            _set(ws, f"J{row}", item.amount, ch)
            _set(ws, f"K{row}", "N.A.", ch)
            row += 1
        if ia_total_item:
            _set(ws, "J28", ia_total_item.amount, ch)
        _set(ws, "K28", "N.A.", ch)

    # Stale gross block number
    from openpyxl.cell.cell import MergedCell
    b38 = ws["B38"]
    if not isinstance(b38, MergedCell) and str(b38.value or "").strip() in {"30991", "30,991"}:
        b38.value = NA_TEXT
        ch.append("Note 12!B38: stale template number cleared")

    for coord in ("A21", "A29", "A34", "A37"):
        cell = ws[coord]
        if not isinstance(cell, MergedCell) and isinstance(cell.value, str) and "Previous Year" in cell.value:
            cell.value = f"Previous Year ({m.prior_year_date})"

    _set(ws, "A41",
         f"The company has not revalued any of its Property, Plant and Equipment for FY {m.reporting_year}.", ch)
    _wipe_sample_text(ws, ch)


# --------------------------------------------------------------------------- #
# CWIP
# --------------------------------------------------------------------------- #

def _write_cwip(wb: "Workbook", m: NoteDataModel, ch: list[str]) -> None:
    ws = _ws(wb, "13.CWIP")
    if not ws:
        return
    yr = m.reporting_date[-4:]
    _set(ws, "B3",
         f"Notes to Financial Statements for the year ended at {m.reporting_date}", ch)
    _set(ws, "F9", f"31.03.{yr}", ch)
    _set(ws, "C9", f"01.04.{int(yr)-1}", ch)
    _set(ws, "B20", f"As at {m.reporting_date}", ch)
    _set(ws, "B26", f"As at {m.prior_year_date} (N.A.)", ch)

    cwip = m.by_note("13")
    for coord in ("B11", "B12", "B13"):
        from openpyxl.cell.cell import MergedCell
        cell = ws[coord]
        if not isinstance(cell, MergedCell):
            if isinstance(cell.value, str) and ("Capital Work" in cell.value or _SAMPLE_RE.search(cell.value)):
                cell.value = NA_TEXT
                ch.append(f"CWIP!{coord}: project name cleared")

    if cwip and cwip.status == "N.A.":
        _set(ws, "B11", "Nil — No CWIP accounts identified in Trial Balance", ch)

    # Clear raw datetime strings
    from openpyxl.cell.cell import MergedCell
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            if isinstance(cell.value, str) and re.match(r"^\d{4}-\d{2}-\d{2}", cell.value):
                cell.value = NA_TEXT
                ch.append(f"CWIP!{cell.coordinate}: raw datetime cleared")

    _wipe_sample_text(ws, ch)


# --------------------------------------------------------------------------- #
# Note (13-20)
# --------------------------------------------------------------------------- #

def _write_note13_20(wb: "Workbook", m: NoteDataModel, ch: list[str]) -> None:
    ws = _ws(wb, "Note (13-20)")
    if not ws:
        return

    # ---- Note 14: investments -------------------------------------- #
    for coord in ("B12", "B13", "B14"):
        from openpyxl.cell.cell import MergedCell
        cell = ws[coord]
        if not isinstance(cell, MergedCell) and isinstance(cell.value, str):
            if any(k in cell.value for k in ("ABCDE", "subsidiary", "19,999", "40,766")):
                cell.value = NA_TEXT
                ch.append(f"Note (13-20)!{coord}: sample investment cleared")

    # ---- Note 17: Trade Receivables -------------------------------- #
    tr = m.by_note("17")
    _set(ws, "H38", f"As at {m.reporting_date}", ch)
    _set(ws, "H48", f"As at {m.prior_year_date} (N.A.)", ch)
    _wipe_numerics(ws, list("CDEFGH"), range(40, 57))
    if tr and tr.total is not None:
        # Write accounts
        row = 44
        for item in [i for i in tr.line_items if isinstance(i.amount, (int, float))]:
            _set(ws, f"H{row}", item.amount, ch)
            _set(ws, f"B{row}", item.label, ch)
            row += 1
        _set(ws, "H46", tr.total, ch)
        _set(ws, "H56", "N.A.", ch)
        # Ageing/classification
        for row_n in (44, 45, 54, 55):
            from openpyxl.cell.cell import MergedCell
            c = ws[f"C{row_n}"]
            if not isinstance(c, MergedCell):
                c.value = NA_TEXT

    # ---- Note 18: Cash & Bank -------------------------------------- #
    cash = m.by_note("18")
    _wipe_numerics(ws, list("EFGH"), range(60, 66))
    if cash:
        row = 62
        for item in [i for i in cash.line_items if isinstance(i.amount, (int, float))]:
            _set(ws, f"F{row}", item.amount, ch)
            _set(ws, f"B{row}", item.label, ch)
            row += 1
        if cash.total:
            _set(ws, "F63", cash.total, ch)
        _set(ws, "G63", "N.A.", ch)

    # ---- Note 19: Short-term loans & advances ---------------------- #
    adv = m.by_note("19")
    _wipe_numerics(ws, list("EFGH"), range(68, 74))
    if adv:
        row = 68
        for item in [i for i in adv.line_items if isinstance(i.amount, (int, float))]:
            _set(ws, f"F{row}", item.amount, ch)
            _set(ws, f"B{row}", item.label, ch)
            row += 1
        if adv.total:
            _set(ws, "F72", adv.total, ch)
        _set(ws, "G72", "N.A.", ch)
    _set(ws, "B74",
         "* Related party advances: Information is not determinable from the Trial Balance "
         "and requires additional management information.", ch)
    _set(ws, "C75", f"As at {m.reporting_date}", ch)
    _set(ws, "D75", "N.A.", ch)
    _set(ws, "E75", f"As at {m.prior_year_date} (N.A.)", ch)

    # ---- Note 20: Other Current Assets ----------------------------- #
    oca = m.by_note("20")
    _wipe_numerics(ws, list("EFGH"), range(82, 90))
    if oca:
        row = 83
        for item in [i for i in oca.line_items if isinstance(i.amount, (int, float))]:
            _set(ws, f"F{row}", item.amount, ch)
            _set(ws, f"B{row}", item.label, ch)
            row += 1
        if oca.total:
            _set(ws, "F88", oca.total, ch)
        _set(ws, "G88", "N.A.", ch)

    _wipe_sample_text(ws, ch)


# --------------------------------------------------------------------------- #
# Note 20-31
# --------------------------------------------------------------------------- #

def _write_note20_31(wb: "Workbook", m: NoteDataModel, ch: list[str]) -> None:
    ws = _ws(wb, "Note 20-31")
    if not ws:
        return
    ry = m.reporting_date

    # Period headers
    for coord, val in [
        ("C64", f"For the period ended {ry}"), ("D64", "N.A."),
        ("C75", f"For the period ended {ry}"), ("D75", "N.A."),
        ("H75", "N.A."), ("H76", "N.A."),
        ("C107", f"For the period ended {ry}"), ("D107", "N.A."),
        ("C121", f"For the period ended {ry}"), ("D121", "N.A."),
        ("C129", f"For the period ended {ry}"), ("D129", "N.A."),
    ]:
        _set(ws, coord, val, ch)

    # ---- Note 21: Revenue ----------------------------------------- #
    rev = m.by_note("21")
    _wipe_numerics(ws, list("CD"), range(7, 13))
    if rev:
        row = 9
        for item in [i for i in rev.line_items if isinstance(i.amount, (int, float))]:
            _set(ws, f"C{row}", item.amount, ch)
            _set(ws, f"B{row}", item.label, ch)
            _set(ws, f"D{row}", "N.A.", ch)
            row += 1
        if rev.total:
            _set(ws, "C11", rev.total, ch)
            _set(ws, "D11", "N.A.", ch)

    # ---- Note 22: Other Income ------------------------------------ #
    oi = m.by_note("22")
    _wipe_numerics(ws, list("CD"), range(15, 24))
    if oi:
        row = 16
        for item in [i for i in oi.line_items if isinstance(i.amount, (int, float))]:
            _set(ws, f"C{row}", item.amount, ch)
            _set(ws, f"B{row}", item.label, ch)
            _set(ws, f"D{row}", "N.A.", ch)
            row += 1
        if oi.total:
            _set(ws, "C23", oi.total, ch)
            _set(ws, "D23", "N.A.", ch)

    # ---- Notes 23-25: material cost / inventory ------------------- #
    _wipe_numerics(ws, list("CD"), range(29, 62))
    _set(ws, "B43", "* Purchase from related party: Information is not determinable from the Trial Balance "
         "and requires additional management information.", ch)

    # ---- Note 26: Employee Benefits ------------------------------- #
    eb = m.by_note("26")
    _wipe_numerics(ws, list("CD"), range(65, 73))
    if eb:
        row = 66
        for item in [i for i in eb.line_items if isinstance(i.amount, (int, float))]:
            _set(ws, f"C{row}", item.amount, ch)
            _set(ws, f"B{row}", item.label, ch)
            _set(ws, f"D{row}", "N.A.", ch)
            row += 1
        if eb.total:
            _set(ws, "C71", eb.total, ch)
            _set(ws, "D71", "N.A.", ch)

    # ---- Note 27/Dep: Other Expenses + Depreciation --------------- #
    dep = m.by_note("dep")
    oe  = m.by_note("27")
    _wipe_numerics(ws, list("CDH"), range(76, 104))

    # Fix wrong depreciation label (template bug: "Installation and Job Work")
    from openpyxl.cell.cell import MergedCell
    b82 = ws["B82"]
    if not isinstance(b82, MergedCell) and isinstance(b82.value, str) and "Installation" in b82.value:
        b82.value = "Depreciation and amortisation expense"
        ch.append("Note 20-31!B82: label corrected to Depreciation")

    if dep and dep.total is not None:
        _set(ws, "C82", dep.total, ch)
        _set(ws, "D82", "N.A.", ch)

    # Write individual other expense accounts
    if oe:
        expense_rows = {
            "audit fee":           77, "audit":              77,
            "bank charge":         78, "bank charges":       78,
            "fees and taxes":      79, "fees & taxes":       79,
            "professional":        80,
            "interest on tds":     81,
            "round":               83,
        }
        placed: set[int] = set()
        row = 77
        for item in [i for i in oe.line_items if isinstance(i.amount, (int, float))]:
            lbl_key = item.label.lower()
            target_row = None
            for key, r in expense_rows.items():
                if key in lbl_key and r not in placed:
                    target_row = r
                    break
            if target_row is None:
                while row in placed or row == 82:
                    row += 1
                    if row > 92:
                        break
                target_row = row

            if target_row <= 92 and target_row != 82:
                _set(ws, f"B{target_row}", item.label, ch)
                _set(ws, f"C{target_row}", item.amount, ch)
                _set(ws, f"D{target_row}", "N.A.", ch)
                placed.add(target_row)
                row = target_row + 1

        if oe.total is not None:
            _set(ws, "C93", oe.total, ch)
            _set(ws, "D93", "N.A.", ch)

    # Prior-year 'N.A.' for H column
    for r in range(75, 104):
        _set(ws, f"H{r}", "N.A.", ch)

    # ---- Note 28: Finance Costs ----------------------------------- #
    fin = m.by_note("28")
    _wipe_numerics(ws, list("CD"), range(108, 117))
    if fin and fin.total:
        row = 109
        for item in [i for i in fin.line_items if isinstance(i.amount, (int, float))]:
            _set(ws, f"C{row}", item.amount, ch)
            _set(ws, f"B{row}", item.label, ch)
            row += 1
        _set(ws, "C116", fin.total, ch)
        _set(ws, "D116", "N.A.", ch)

    # ---- Note 29: Current Tax ------------------------------------- #
    tax = m.by_note("29")
    _wipe_numerics(ws, list("CD"), range(122, 126))
    if tax and tax.total:
        _set(ws, "C123", tax.total, ch)
        _set(ws, "C124", tax.total, ch)
        _set(ws, "D124", "N.A.", ch)

    # ---- Note 30: EPS --------------------------------------------- #
    eps = m.by_note("30")
    _set(ws, "C129", f"For the period ended {ry}", ch)
    _wipe_numerics(ws, list("CD"), range(132, 141))
    if eps:
        pat_item = next((i for i in eps.line_items if "profit" in i.label.lower()), None)
        if pat_item and isinstance(pat_item.amount, (int, float)):
            _set(ws, "C133", pat_item.amount, ch)
        face_item = next((i for i in eps.line_items if "face value" in i.label.lower()), None)
        if face_item:
            _set(ws, "C137", face_item.amount if isinstance(face_item.amount, (int, float)) else "₹10", ch)
        for coord in ("C134", "C135", "C136", "C139", "C140"):
            _set(ws, coord, NA_TEXT, ch)

    # ---- Note 31: Previous Year figures --------------------------- #
    _set(ws, "B143",
         "Previous year comparative figures are not available as no prior-year Trial Balance was provided.", ch)

    _wipe_sample_text(ws, ch)


# --------------------------------------------------------------------------- #
# EBP
# --------------------------------------------------------------------------- #

def _write_ebp(wb: "Workbook", m: NoteDataModel, ch: list[str]) -> None:
    ws = _ws(wb, "EBP")
    if not ws:
        return
    _set(ws, "A2",
         f"Notes to Financial Statements for the period ended at {m.reporting_date}", ch)
    _set(ws, "F13", m.reporting_date, ch)
    _set(ws, "G13", "N.A.", ch)
    _set(ws, "F12", f"For the year ended {m.reporting_date}", ch)

    # Actuarial table values — all N.A.
    actuarial_cells = {
        "E155", "E156", "E157", "E158", "E159", "E160", "E161", "E162", "E163", "E165",
    }
    from openpyxl.cell.cell import MergedCell
    for coord in actuarial_cells:
        cell = ws[coord]
        if not isinstance(cell, MergedCell) and cell.value not in (None, "N.A.", NA_TEXT):
            cell.value = NA_TEXT
            ch.append(f"EBP!{coord}: actuarial sample cleared")

    for coord in ("E36", "E53", "E70", "E84", "E98", "E107", "E119", "E137", "E154", "E168"):
        cell = ws[coord]
        if not isinstance(cell, MergedCell):
            if cell.value and re.search(r"2023|2022", str(cell.value)):
                cell.value = NA_TEXT
            elif cell.value and re.search(r"2024", str(cell.value)):
                cell.value = f"{m.prior_year_date} (N.A.)"

    # Wipe all numeric actuarial values
    for row in range(34, 180):
        for col in ("C", "D", "E"):
            from openpyxl.cell.cell import MergedCell as MC
            cell = ws[f"{col}{row}"]
            if not isinstance(cell, MC) and isinstance(cell.value, (int, float)):
                cell.value = None
    _wipe_sample_text(ws, ch)


# --------------------------------------------------------------------------- #
# Other Notes
# --------------------------------------------------------------------------- #

def _write_other_notes(wb: "Workbook", m: NoteDataModel, ch: list[str]) -> None:
    ws = _ws(wb, "Other Notes")
    if not ws:
        return
    yr  = m.reporting_date[-4:]
    ry  = m.reporting_date
    py  = m.prior_year_date

    # MSME column headers
    _set(ws, "E8", f"31st March, {yr}", ch)
    _set(ws, "F8", f"31st March, {int(yr)-1} (N.A.)", ch)

    # Sec 186
    _set(ws, "E21", f"As at {ry}", ch)
    _set(ws, "G21", "N.A.", ch)
    _set(ws, "G22", "N.A.", ch); _set(ws, "H22", "N.A.", ch)
    _set(ws, "G28", "N.A.", ch); _set(ws, "H28", "N.A.", ch)
    from openpyxl.cell.cell import MergedCell
    b29 = ws["B29"]
    if not isinstance(b29, MergedCell) and isinstance(b29.value, str) and "XYZ Limited" in b29.value:
        b29.value = NA_TEXT
        ch.append("Other Notes!B29: sample entity cleared")

    # Solvency ratios
    _set(ws, "F33", f"For the year ended {ry}", ch)
    _set(ws, "G33", "N.A.", ch)
    _set(ws, "H33", "N.A.", ch)
    for row in range(34, 46):
        for col in ("F", "G", "H"):
            cell = ws[f"{col}{row}"]
            if not isinstance(cell, MergedCell) and isinstance(cell.value, (int, float)):
                cell.value = NA_TEXT

    # Contingent liabilities
    _set(ws, "F55", f"As at {ry}", ch)
    _set(ws, "G55", "N.A.", ch)
    for row in (57, 58, 61, 64):
        for col in ("E", "F"):
            cell = ws[f"{col}{row}"]
            if not isinstance(cell, MergedCell) and isinstance(cell.value, (int, float)):
                cell.value = NA_TEXT

    # Related parties — clear placeholder names
    for coord in ("B70", "B71", "B72", "B73", "B74"):
        cell = ws[coord]
        if not isinstance(cell, MergedCell) and isinstance(cell.value, str):
            if any(s in cell.value for s in ("…", "........", "Punit", "XYZ", "Director")):
                cell.value = NA_TEXT
                ch.append(f"Other Notes!{coord}: RP placeholder cleared")
    _set(ws, "F79", f"For the Year ended\n{ry}", ch)
    _set(ws, "G79", "N.A.", ch)
    for row in range(80, 107):
        for col in ("F", "G"):
            cell = ws[f"{col}{row}"]
            if not isinstance(cell, MergedCell) and isinstance(cell.value, (int, float)):
                cell.value = NA_TEXT

    # Segment
    _set(ws, "B109", "Segment information is not available from the Trial Balance.", ch)

    # CSR
    _set(ws, "B112",
         "CSR applicability and related disclosure require additional company information.", ch)

    # Events after BS date
    _set(ws, "B141",
         "Information regarding events after the reporting date cannot be determined from "
         "the Trial Balance and requires additional management information.", ch)

    # Note 44 regulatory — fix years and assertions
    for row in range(145, 170):
        cell = ws[f"B{row}"]
        if isinstance(cell.value, str):
            updated = cell.value\
                .replace("March 31, 2025", f"March 31, {yr}")\
                .replace("2024-25", m.reporting_year)
            if updated != cell.value:
                cell.value = updated
    _set(ws, "B155",
         "Disclosure regarding loans/advances to promoters, directors, KMP and related parties "
         "requires management confirmation; not determinable from the Trial Balance alone.", ch)
    _set(ws, "F156", m.reporting_year, ch)
    _set(ws, "D156", "N.A.", ch); _set(ws, "D157", "N.A.", ch); _set(ws, "G157", "N.A.", ch)
    for row in (158, 159, 160, 161):
        _set(ws, f"F{row}", NA_TEXT, ch)
        _set(ws, f"G{row}", "N.A.", ch)

    _wipe_sample_text(ws, ch)


# --------------------------------------------------------------------------- #
# Stock Reconciliation
# --------------------------------------------------------------------------- #

def _write_stock_recon(wb: "Workbook", m: NoteDataModel, ch: list[str]) -> None:
    ws_name = next((n for n in wb.sheetnames if "Stock" in n), None)
    if not ws_name:
        return
    ws = wb[ws_name]
    _set(ws, "E5", f"For the year ended {m.reporting_date}", ch)
    _set(ws, "I5", f"For the year ended {m.prior_year_date} (N.A.)", ch)

    from openpyxl.cell.cell import MergedCell
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            if isinstance(cell.value, str) and re.match(r"^\d{4}-\d{2}-\d{2}", cell.value):
                cell.value = NA_TEXT
                ch.append(f"StockRecon!{cell.coordinate}: raw datetime cleared")
            elif isinstance(cell.value, str) and _SAMPLE_RE.search(cell.value):
                cell.value = NA_TEXT

    cell = ws["B22"]
    if isinstance(cell.value, str):
        updated = cell.value.replace("March 31, 2025", f"March 31, {m.reporting_date[-4:]}")
        if updated != cell.value:
            cell.value = updated

    if isinstance(ws["B23"].value, str) and ws["B23"].value.strip() in ("\\", "\\'"):
        ws["B23"].value = ""

    for row in range(7, 22):
        for col in ("E", "F", "G", "H", "I", "J", "K", "L"):
            cell = ws[f"{col}{row}"]
            if not isinstance(cell, MergedCell) and isinstance(cell.value, (int, float)):
                cell.value = NA_TEXT


# --------------------------------------------------------------------------- #
# Borrowing
# --------------------------------------------------------------------------- #

def _write_borrowing(wb: "Workbook", m: NoteDataModel, ch: list[str]) -> None:
    ws = _ws(wb, "Borrowing")
    if not ws:
        return
    yr = m.reporting_date[-4:]
    _set(ws, "E5", f"As at {m.reporting_date}", ch)
    _set(ws, "G5", f"As at {m.prior_year_date} (N.A.)", ch)

    borrow = m.by_note("5")
    row = 8
    for item in (borrow.line_items if borrow else []):
        if item.review_reason and isinstance(item.amount, (int, float)):
            _set(ws, f"C{row}", item.label, ch)
            _set(ws, f"E{row}", item.amount, ch)
            _set(ws, f"G{row}", "N.A.", ch)
            row += 1

    from openpyxl.cell.cell import MergedCell
    for row_n in range(5, 32):
        for col in ("G", "H"):
            cell = ws[f"{col}{row_n}"]
            if not isinstance(cell, MergedCell) and cell.value not in (None, "N.A.") and isinstance(cell.value, (int, float)):
                cell.value = "N.A."


# --------------------------------------------------------------------------- #
# EPS
# --------------------------------------------------------------------------- #

def _write_eps(wb: "Workbook", m: NoteDataModel, ch: list[str]) -> None:
    ws = _ws(wb, "EPS")
    if not ws:
        return
    _set(ws, "C5", f"For the period ended at {m.reporting_date}", ch)
    _set(ws, "D5", "N.A.", ch)

    eps = m.by_note("30")
    _wipe_numerics(ws, list("CD"), range(9, 17))
    if eps:
        pat_item = next((i for i in eps.line_items if "profit" in i.label.lower()), None)
        if pat_item and isinstance(pat_item.amount, (int, float)):
            _set(ws, "C9", pat_item.amount, ch)
        _set(ws, "C13", "₹10", ch)
        for coord in ("C10", "C11", "C12", "C15", "C16"):
            _set(ws, coord, NA_TEXT, ch)
