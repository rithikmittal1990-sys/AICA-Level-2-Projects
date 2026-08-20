"""Remove template/sample financial data before writing trial balance values."""

from __future__ import annotations

import re
from typing import Iterable

from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.excel.template_manager import canonical_sheet_name

EXTERNAL_WORKBOOK_RE = re.compile(r"\[\d+\]")
COMPARATIVE_COLUMNS = {"D", "E", "F", "G", "H", "I", "J", "K", "L"}
AMOUNT_COLUMNS = set("CDEFGHIJKL")
CLEARABLE_SHEETS = {
    "Note 12 (PPE)", "Note (13-20)", "Note 20-31", "EPS", "Ratio_working",
    "Other Notes", "Borrowing", "EBP", "13.CWIP", "NOTE (4-12)",
}
LABEL_COLUMNS = {"A", "B"}
EXCLUDED_SHEETS = {"Cash Flow", "Validation"}
KNOWN_TEMPLATE_NUMBERS = {
    9975360, 8312800, 166256, 1496304, 58538616, 80457755, 83128000, 3314240,
    119575580, 9652174, 4054404, 242430740, 16000000,
}


def clear_template_financial_data(
    workbook: Workbook,
    *,
    exclude_sheets: Iterable[str] = EXCLUDED_SHEETS,
) -> list[dict[str, str]]:
    """Clear numeric values and external-reference formulas from the template copy."""
    cleared: list[dict[str, str]] = []
    excluded = {canonical_sheet_name(name) for name in exclude_sheets}
    for sheet in workbook.worksheets:
        if canonical_sheet_name(sheet.title) in excluded:
            continue
        aggressive = canonical_sheet_name(sheet.title) in {canonical_sheet_name(s) for s in CLEARABLE_SHEETS}
        cleared.extend(_clear_sheet(sheet, aggressive=aggressive))
    return cleared


def _clear_sheet(sheet: Worksheet, *, aggressive: bool = False) -> list[dict[str, str]]:
    cleared: list[dict[str, str]] = []
    for row in sheet.iter_rows():
        for cell in row:
            row_idx = cell.row
            col = re.sub(r"\d+", "", cell.coordinate)
            value = cell.value
            if col in LABEL_COLUMNS:
                if not (
                    aggressive
                    and isinstance(value, (int, float))
                    and not isinstance(value, bool)
                    and abs(float(value)) >= 100000
                ):
                    continue
            if cell.row <= 4 and col in {"A", "B", "C"}:
                continue
            if isinstance(value, str) and value.startswith("="):
                if EXTERNAL_WORKBOOK_RE.search(value):
                    cell.value = None
                    cleared.append({"sheet": sheet.title, "cell": cell.coordinate, "action": "cleared_external_formula"})
                elif col in AMOUNT_COLUMNS and row_idx > 6:
                    cell.value = None
                    cleared.append({"sheet": sheet.title, "cell": cell.coordinate, "action": "cleared_amount_formula"})
                continue
            if isinstance(value, (int, float)) and not isinstance(value, bool):
                if col in AMOUNT_COLUMNS and row_idx > 6:
                    cell.value = None
                    cleared.append({"sheet": sheet.title, "cell": cell.coordinate, "action": "cleared_value"})
                elif aggressive and row_idx > 3 and abs(float(value)) > 0.01:
                    if col in LABEL_COLUMNS and isinstance(value, (int, float)) and abs(float(value)) < 100000:
                        continue
                    if col not in LABEL_COLUMNS or isinstance(value, (int, float)):
                        cell.value = None if isinstance(value, (int, float)) else cell.value
                        if isinstance(value, (int, float)):
                            cleared.append({"sheet": sheet.title, "cell": cell.coordinate, "action": "cleared_value"})
    return cleared


def clear_comparative_columns(workbook: Workbook) -> None:
    """Previous-year columns must not retain template sample numbers."""
    for sheet in workbook.worksheets:
        if canonical_sheet_name(sheet.title) in EXCLUDED_SHEETS:
            continue
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row or 1):
            for cell in row:
                col = re.sub(r"\d+", "", cell.coordinate)
                if col in {"G", "H", "L", "D"} and cell.row > 6:
                    if isinstance(cell.value, (int, float)):
                        cell.value = None
                    elif isinstance(cell.value, str) and cell.value.strip().upper() not in {"", "N.A.", "NA"}:
                        if not cell.value.startswith("="):
                            cell.value = "N.A."


_PLACEHOLDER_STRINGS = re.compile(
    r"\b(XYZ Limited(?! Private)|ABCDE|Punit Kumar Sarda|Object Clause"
    r"|99,75,360|9,975,360|8,312,800|1,662,560|1,496,304"
    r"|\.{4,}\s*(Director|Relative)|Director's Relative"
    r"|sample director|sample shareholder|sample company|sample related party)\b",
    re.IGNORECASE,
)
_OLD_DATE_RE = re.compile(
    r"\b(31\s*March\s*202[34]|31st March,?\s*202[34]|March 31,?\s*202[34]"
    r"|2023-24|2022-23|FY\s*202[34]-\d\d)\b",
    re.IGNORECASE,
)


def full_leakage_scan(workbook: Workbook, company_name: str | None = None) -> list[dict[str, str]]:
    """
    Full textual + numeric leakage scan across ALL sheets (including hidden).
    Returns a list of {sheet, cell, type, value} dicts.
    """
    issues: list[dict[str, str]] = []
    skip = {"Validation"}
    known_nums = KNOWN_TEMPLATE_NUMBERS | {99753600, 83128000, 16625600, 14963040}

    for sheet in workbook.worksheets:
        if sheet.title in skip:
            continue
        for row in sheet.iter_rows():
            for cell in row:
                val = cell.value
                if val is None:
                    continue
                coord = f"{sheet.title}!{cell.coordinate}"

                if isinstance(val, (int, float)) and not isinstance(val, bool):
                    fv = float(val)
                    iv = int(fv) if fv == int(fv) else fv
                    if iv in known_nums:
                        issues.append({"sheet": sheet.title, "cell": cell.coordinate,
                                       "type": "KNOWN_TEMPLATE_NUMBER", "value": str(val)})
                    elif fv >= 10_000_000:
                        issues.append({"sheet": sheet.title, "cell": cell.coordinate,
                                       "type": "LARGE_NUMBER_SUSPECT", "value": str(val)})

                elif isinstance(val, str):
                    if _PLACEHOLDER_STRINGS.search(val):
                        issues.append({"sheet": sheet.title, "cell": cell.coordinate,
                                       "type": "PLACEHOLDER_NAME", "value": val[:100]})
                    if _OLD_DATE_RE.search(val) and "N.A." not in val:
                        # Allow prior-year column headers (e.g. "As at 31st March, 2025")
                        if not re.search(r"(?:As at|as at|prior|previous)", val, re.IGNORECASE):
                            issues.append({"sheet": sheet.title, "cell": cell.coordinate,
                                           "type": "OLD_DATE_IN_TEXT", "value": val[:100]})
                    if re.search(r"Rs\.?\s*[\d,]{8,}", val) and re.search(r"[A-Za-z]", val):
                        issues.append({"sheet": sheet.title, "cell": cell.coordinate,
                                       "type": "HARDCODED_RUPEE_AMOUNT", "value": val[:100]})

    return issues


def detect_template_leakage(workbook: Workbook) -> list[dict[str, str | float]]:
    """Find surviving values that match known ICAI sample magnitudes."""
    leaks: list[dict[str, str | float]] = []
    for sheet in workbook.worksheets:
        if canonical_sheet_name(sheet.title) in EXCLUDED_SHEETS | {"Validation"}:
            continue
        for row in sheet.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, (int, float)) and not isinstance(value, bool):
                    iv = int(value) if float(value).is_integer() else float(value)
                    if iv in KNOWN_TEMPLATE_NUMBERS or abs(float(value)) >= 10_000_000:
                        leaks.append(
                            {
                                "sheet": sheet.title,
                                "cell": cell.coordinate,
                                "value": float(value),
                            }
                        )
    return leaks
