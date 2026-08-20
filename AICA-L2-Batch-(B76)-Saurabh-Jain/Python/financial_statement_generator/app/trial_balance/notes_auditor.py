"""
Textual audit of all Notes sheets in the generated workbook.

Removes / replaces:
  - Sample company names, person names, subsidiary names
  - Wrong year dates (template says 2024-25 / March 31 2025; fix to actual period)
  - Old comparative column headers that no longer apply
  - Raw Python datetime objects rendered as strings
  - Hardcoded related-party rupee amounts from sample data
  - Boilerplate management/legal assertions that cannot be verified from the TB
  - Empty placeholder rows containing only '0', '-', '\\', or whitespace

Does NOT touch:
  - Any numeric cell that was written by the classifier / statement builder
  - Structural labels (row headings, section headings)
  - Correctly-written N.A. markers already in place
"""

from __future__ import annotations

import re
from datetime import datetime
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from openpyxl import Workbook
    from openpyxl.worksheet.worksheet import Worksheet

NA = "Information not available from Trial Balance"
_BLANK = ""

# --------------------------------------------------------------------------- #
# Patterns that identify template / sample text that must be cleared
# --------------------------------------------------------------------------- #

# Sample company / person names that originate from the template
_SAMPLE_ENTITY_RE = re.compile(
    r"\b(ABC India Limited|ABC India|ABCDE|Punit Kumar Sarda|XYZ Limited(?! Private))\b",
    re.IGNORECASE,
)

# Hardcoded rupee figures from sample data (e.g. "Rs. 1,34,97,237" or "Rs. 6,70,56,487")
_SAMPLE_RUPEE_RE = re.compile(
    r"Rs\.?\s*[\d,]+(?:\.\d+)?\s*(?:\(P\.?Y\.?\s*Rs\.?\s*[\d,]+(?:\.\d+)?\))?"
)

# Raw Python datetime string rendered into a cell  e.g. "2024-06-01 00:00:00"
_DATETIME_STR_RE = re.compile(r"^\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2}$")

# Stray lone backslash  (appears in Stock Reconciliation)
_STRAY_BACKSLASH_RE = re.compile(r"^\\+$")

# A cell whose entire content is just '-' or '—' separators (not a number)
_DASH_ONLY_RE = re.compile(r"^[-–—\s]+$")

# Old year strings in column/row headers
_OLD_YEAR_PATTERNS = [
    ("March 31, 2025", "March 31, 2026"),
    ("March 31, 2024", "March 31, 2025"),
    ("31st March, 2025", "31st March, 2026"),
    ("31st March, 2024", "31st March, 2025"),
    ("31st March 2025", "31st March 2026"),
    ("31st March 2024", "31st March 2025"),
    ("31st March, 2023", "N.A."),
    ("31st March 2023", "N.A."),
    ("31.03.2025", "31.03.2026"),
    ("31.03.2024", "31.03.2025"),
    ("01.04.2024", "01.04.2025"),
    ("01.04.2023", "01.04.2024"),
    ("FY 2024-25", "FY 2025-26"),
    ("FY 2023-24", "FY 2024-25"),
    ("for the year ended March 31, 2025", "for the year ended March 31, 2026"),
    ("for the year ended at 31st March 2025", "for the year ended at 31st March 2026"),
    ("for the year ended at March 31, 2025", "for the year ended at March 31, 2026"),
    ("For the year ended 31st March 2025", "For the year ended 31st March 2026"),
    ("For the period ended at March 31, 2025", "For the period ended at March 31, 2026"),
    ("for the period ended at March 31, 2025", "for the period ended at March 31, 2026"),
    ("for the period ended at 31st March 2025", "for the period ended at 31st March 2026"),
    ("31-Mar-2025", "31-Mar-2026"),
    ("31-Mar-2024", "31-Mar-2025"),
    ("2024-25", "2025-26"),
    ("2023-24", "2024-25"),
    ("2022-23", "N.A."),
    ("2021-22", "N.A."),
    ("2020-21", "N.A."),
]

# Assertions that refer to management representations not derivable from TB alone.
# These cells will be replaced with NA.
_UNSUPPORTED_ASSERTIONS: dict[str, dict[str, str]] = {
    "Note 1-2": {
        # Note 2 accounting policies that refer to specific company operations not in TB
        "B34": NA,   # Income Taxes (generic boilerplate — retain but don't touch numbers)
    },
    "Other Notes": {
        # Related party names in Note 36 are placeholders
        "B70": NA,
        "B71": NA,
        "B72": NA,
        "B73": NA,
        "B74": NA,
    },
}

# Stray sample/hardcoded number cells that aren't financial calculations
_STRAY_NUMBERS: dict[str, set[str]] = {
    "Note 12 (PPE)": {"B38"},   # "30991" — stale gross block total from template
}

# Cells containing only placeholder zeros in comparative columns
_ZERO_PLACEHOLDER_CELLS: dict[str, set[str]] = {
    "Note 3 (Share Capital)": {"O51", "O53", "M55", "O55", "O57", "O66", "O68", "O70", "O72"},
}

# Share capital descriptive text to replace (authorised / issued numbers are sample)
_SHARE_CAPITAL_SAMPLE_CELLS: dict[str, str] = {
    "B10": "Authorised share capital — Information not available from Trial Balance",
    "B12": "Issued, subscribed and fully paid up equity shares",
}

# Note 20-31 B82: label says "Installation and Job Work Expenses" but the value is depreciation
_WRONG_LABELS: dict[str, dict[str, str]] = {
    "Note 20-31": {
        "B82": "Depreciation and amortisation expense",
        # Remove hardcoded related party purchase text
        "B43": "* Purchase from related party: Information not available from Trial Balance",
        "B97": "* Trade payable related party details: Information not available from Trial Balance",
    },
    "NOTE (4-12)": {
        "B97": "* Trade Payable related party details: Information not available from Trial Balance",
        "B31": "Refer borrowing schedule",
        "B72": "Refer borrowing schedule",
    },
    "Note (13-20)": {
        "B13": "Information not available from Trial Balance",
        "B14": "Information not available from Trial Balance",
        "B29": "Information not available from Trial Balance",
    },
    "Other Notes": {
        "B109": (
            "The Company primarily operates in the business segment indicated by its object clause. "
            "Segment information is not separately derivable from the uploaded Trial Balance."
        ),
        "B112": (
            "Applicability of Section 135 of the Companies Act, 2013 requires assessment by management; "
            "not determinable from Trial Balance alone."
        ),
        "B155": (
            "The Company has/has not granted loans and advances in the nature of loans to promoters, directors, "
            "KMP and other related parties — details are not available from Trial Balance."
        ),
    },
    "EBP": {
        # Actuarial assumptions are sample data; mark entire note as NA
    },
    "13.CWIP": {
        "B11": NA,
        "B12": NA,
        "B13": NA,
    },
}

# Cells in EBP that contain specific actuarial numbers from sample data — clear them
_EBP_ACTUARIAL_CELLS = {"E155", "E156", "E157", "E158", "E159", "E160",
                         "E161", "E162", "E163", "E165"}


def _apply_year_corrections(text: str) -> str:
    """Replace old year strings in text with corrected ones."""
    for old, new in _OLD_YEAR_PATTERNS:
        text = text.replace(old, new)
    return text


def _is_numeric_value(cell_value) -> bool:
    """Return True if cell holds a genuine numeric result (not a stray sample number)."""
    return isinstance(cell_value, (int, float))


def _clear_if_sample_entity(ws: "Worksheet", coord: str) -> None:
    cell = ws[coord]
    val = cell.value
    if val is None:
        return
    s = str(val)
    if _SAMPLE_ENTITY_RE.search(s):
        cell.value = NA


def audit_notes_sheets(
    workbook: "Workbook",
    company_name: str | None,
    period_label: str | None,
    year_end: str = "31st March, 2026",
    prior_year_end: str = "31st March, 2025",
) -> dict[str, list[str]]:
    """
    Walk every Notes sheet and apply all textual corrections in-place.

    Returns a dict {sheet_name: [list of change descriptions]} for logging.
    """
    changes: dict[str, list[str]] = {}

    for sheet_name in workbook.sheetnames:
        if sheet_name in ("BS PnL", "Validation"):
            continue
        ws = workbook[sheet_name]
        sheet_changes: list[str] = []

        for row in ws.iter_rows():
            for cell in row:
                val = cell.value
                if val is None:
                    continue

                coord = cell.coordinate

                # ------------------------------------------------------------------ #
                # 1. Raw Python datetime objects rendered as strings
                # ------------------------------------------------------------------ #
                if isinstance(val, datetime):
                    cell.value = NA
                    sheet_changes.append(f"{coord}: datetime object → {NA}")
                    continue

                s = str(val)

                if _DATETIME_STR_RE.match(s):
                    cell.value = NA
                    sheet_changes.append(f"{coord}: raw datetime string → {NA}")
                    continue

                # ------------------------------------------------------------------ #
                # 2. Stray lone backslash placeholder
                # ------------------------------------------------------------------ #
                if _STRAY_BACKSLASH_RE.match(s):
                    cell.value = _BLANK
                    sheet_changes.append(f"{coord}: stray backslash cleared")
                    continue

                # ------------------------------------------------------------------ #
                # 3. Year corrections (string cells only — never touch numeric cells)
                # ------------------------------------------------------------------ #
                if not _is_numeric_value(val):
                    corrected = _apply_year_corrections(s)
                    if corrected != s:
                        cell.value = corrected
                        sheet_changes.append(f"{coord}: year corrected '{s[:60]}' → '{corrected[:60]}'")
                        s = corrected

                # ------------------------------------------------------------------ #
                # 4. Sample entity names inside text
                # ------------------------------------------------------------------ #
                if not _is_numeric_value(val) and _SAMPLE_ENTITY_RE.search(s):
                    cleaned = _SAMPLE_ENTITY_RE.sub(NA, s)
                    cell.value = cleaned
                    sheet_changes.append(f"{coord}: sample entity removed")
                    s = cleaned

                # ------------------------------------------------------------------ #
                # 5. Hardcoded rupee amounts in narrative text (sample data)
                # ------------------------------------------------------------------ #
                if not _is_numeric_value(val) and _SAMPLE_RUPEE_RE.search(s):
                    # Only clear if the entire cell is a narrative note (has letters around ₹)
                    if re.search(r"[A-Za-z]", s):
                        cleaned = _SAMPLE_RUPEE_RE.sub(
                            "Information not available from Trial Balance", s
                        )
                        cell.value = cleaned
                        sheet_changes.append(f"{coord}: hardcoded rupee amount removed")
                        s = cleaned

        # ---------------------------------------------------------------------- #
        # Sheet-specific targeted fixes
        # ---------------------------------------------------------------------- #
        _fix_sheet_specific(ws, sheet_name, sheet_changes)

        changes[sheet_name] = sheet_changes

    return changes


def _fix_sheet_specific(ws: "Worksheet", sheet_name: str, changes: list[str]) -> None:
    """Apply fixes that are unique to specific sheets."""

    # ------------------------------------------------------------------ #
    # Note 1-2
    # ------------------------------------------------------------------ #
    if sheet_name == "Note 1-2":
        # A3 heading: year-end text (already fixed by year loop, but double-check)
        _replace_text(ws, "A3",
                      "Notes forming part of the Financial Statements for the year ended March 31, 2025",
                      "Notes forming part of the Financial Statements for the year ended 31st March, 2026",
                      changes)
        # Remove repeated section number "2.1" at A77 (duplicate of A15)
        if ws["A77"].value == "2.1":
            ws["A77"].value = "2.9"
            changes.append("A77: duplicate section number fixed to 2.9")

    # ------------------------------------------------------------------ #
    # Note 3 (Share Capital)
    # ------------------------------------------------------------------ #
    if sheet_name == "Note 3 (Share Capital)":
        # Authorised capital row — sample data "1,60,00,000 Equity Shares of Rs 10 each"
        _clear_sample_text(ws, "B10", "1,60,00,000 Equity Shares of Rs 10 each",
                            "Equity shares — authorised capital not separately stated in Trial Balance", changes)
        # Issued capital row — sample data "99,75,360 Equity Shares of Rs 10 each"
        _clear_sample_text(ws, "B12", "99,75,360 Equity Shares of Rs 10 each",
                            "Equity shares of Rs 10 each — quantity from Trial Balance", changes)
        # Column headers for current/prior year
        for coord, replacement in [
            ("F7", "As at 31st March, 2026"),
            ("J7", "As at 31st March, 2025"),
            ("F18", "As at 31st March, 2026"),
            ("J18", "As at 31st March, 2025"),
            ("F48", "As at 31st March, 2026"),
            ("K48", "As at 31st March, 2025"),
            ("F63", "As at 31st March, 2026"),
            ("K63", "As at 31st March, 2025"),
        ]:
            _force_set(ws, coord, replacement, changes)
        # Five-year history table year headers (F37=current, J37=N.A. for 2022-23 etc.)
        for coord in ("F37", "J37", "N37"):
            cell = ws[coord]
            if cell.value and str(cell.value).strip() in {"2024-25", "2022-23", "2020-21"}:
                val = str(cell.value).strip()
                replacement = {"2024-25": "2025-26", "2022-23": "N.A.", "2020-21": "N.A."}.get(val, val)
                cell.value = replacement
                changes.append(f"{coord}: year corrected {val} → {replacement}")
        # Placeholder zero cells in comparative columns
        for coord in {"O51", "O53", "M55", "O55", "O57", "O66", "O68", "O70", "O72"}:
            cell = ws[coord]
            if cell.value is not None and str(cell.value).strip() == "0":
                cell.value = "N.A."
                changes.append(f"{coord}: placeholder 0 → N.A.")
        # Terms/rights — remove hardcoded dividend text if it mentions specific figures
        # (B28 is generic boilerplate text — acceptable to keep, no sample numbers)

    # ------------------------------------------------------------------ #
    # NOTE (4-12)
    # ------------------------------------------------------------------ #
    if sheet_name == "NOTE (4-12)":
        # Heading
        _replace_text(ws, "A3",
                      "Notes to financial statements for the year ended at 31st March 2025",
                      "Notes to financial statements for the year ended at 31st March 2026",
                      changes)
        # Trade payable related party footnote with hardcoded sample rupee amounts — already handled
        # by _SAMPLE_RUPEE_RE above. But double-check:
        _clear_sample_text(ws, "B97",
                            "* Trade Payable include payable to related party Rs. 1,34,97,237 (P.Y. Rs. 43,56,174)",
                            "* Trade payable related party details: Information not available from Trial Balance",
                            changes)
        # Deferred tax column headers
        for coord, repl in [("F35", "As at 31st March, 2026"), ("F40", "As at 31st March, 2025")]:
            _force_set(ws, coord, repl, changes)

    # ------------------------------------------------------------------ #
    # Note 12 (PPE)
    # ------------------------------------------------------------------ #
    if sheet_name == "Note 12 (PPE)":
        # Stray gross block total from template (not from TB)
        cell = ws["B38"]
        if cell.value is not None and str(cell.value).strip() == "30991":
            cell.value = NA
            changes.append("B38: stale template gross-block number cleared")
        # Revaluation note: fix year
        _replace_text(ws, "A41",
                      "FY 2024-25",
                      "FY 2025-26",
                      changes)
        # Column date headers
        for coord, repl in [
            ("B8", "ON 01.04.2025"), ("E8", "31.03.2026"),
            ("F8", "31.03.2025"), ("J8", "ON 31.03.2026"), ("K8", "ON 31.03.2025"),
        ]:
            cell = ws[coord]
            if cell.value and str(cell.value).strip() in {
                "ON 01.04.2024", "31.03.2025", "31.03.2024", "ON 31.03.2025", "ON 31.03.2024"
            }:
                cell.value = repl
                changes.append(f"{coord}: date corrected → {repl}")
        # "Previous Year" row labels — make explicit
        for coord in ("A21", "A29", "A34", "A37"):
            cell = ws[coord]
            if cell.value and str(cell.value).strip() == "Previous Year":
                cell.value = "Previous Year (31st March, 2025)"
                changes.append(f"{coord}: 'Previous Year' qualified with date")

    # ------------------------------------------------------------------ #
    # 13.CWIP
    # ------------------------------------------------------------------ #
    if sheet_name == "13.CWIP":
        _replace_text(ws, "B3",
                      "Notes to Financial Statements for the year ended at March 31, 2025",
                      "Notes to Financial Statements for the year ended at March 31, 2026",
                      changes)
        # Column period headers
        _force_set(ws, "F9", "31.03.2026", changes)
        _force_set(ws, "C9", "ON 01.04.2025", changes)
        # CWIP project names — sample placeholders
        for coord in ("B11", "B12", "B13"):
            cell = ws[coord]
            if cell.value and str(cell.value).strip() in {
                "Capital Work In Progress-Computer Software",
                "Capital Work In Progress-Gabion Wall",
                "Capital Work In Progress-Water Tank",
            }:
                cell.value = NA
                changes.append(f"{coord}: sample CWIP project name cleared")
        # Aging schedule period labels
        _force_set(ws, "B20", "As at 31st March 2026", changes)
        _force_set(ws, "B26", "As at 31st March 2025", changes)

    # ------------------------------------------------------------------ #
    # Note (13-20)
    # ------------------------------------------------------------------ #
    if sheet_name == "Note (13-20)":
        # Non-current investments — ABCDE subsidiary (already caught by entity regex)
        # but explicitly clear descriptive rows
        for coord in ("B12", "B13", "B14"):
            cell = ws[coord]
            if cell.value and (
                "ABCDE" in str(cell.value) or
                "subsidiary" in str(cell.value).lower() or
                "19,999" in str(cell.value) or
                "40,766" in str(cell.value)
            ):
                cell.value = NA
                changes.append(f"{coord}: sample subsidiary holding description cleared")
        # Related party advances footnote — B74 onward
        _replace_text(ws, "B74",
                      "*Loan & Advances to Related Party",
                      "* Loan & Advances to Related Party: Information not available from Trial Balance",
                      changes)
        # Column headers
        _force_set(ws, "C75", "As at 31st March, 2026", changes)
        _force_set(ws, "E75", "As at 31st March, 2025", changes)
        # Aging headers
        for coord in ("H38", "H48"):
            cell = ws[coord]
            if cell.value == "N.A.":
                cell.value = "As at 31st March, 2025 N.A."
                changes.append(f"{coord}: prior year aging header marked N.A.")

    # ------------------------------------------------------------------ #
    # Note 20-31
    # ------------------------------------------------------------------ #
    if sheet_name == "Note 20-31":
        # B82 wrong label (depreciation written as "Installation and Job Work Expenses")
        cell = ws["B82"]
        if cell.value and "Installation and Job Work" in str(cell.value):
            cell.value = "Depreciation and amortisation expense"
            changes.append("B82: wrong label 'Installation and Job Work Expenses' → 'Depreciation and amortisation expense'")
        # Purchase from related party footnote (B43)
        _clear_sample_text(
            ws, "B43",
            "* Purchase form related party Rs. 6,70,56,487 (PY Rs. 6,09,94,089)",
            "* Purchase from related party: Information not available from Trial Balance",
            changes,
        )

    # ------------------------------------------------------------------ #
    # EBP (Employee Benefit Plans)
    # ------------------------------------------------------------------ #
    if sheet_name == "EBP":
        _replace_text(ws, "A2",
                      "Notes to Financial Statements for the period ended at March 31, 2025",
                      "Notes to Financial Statements for the period ended at March 31, 2026",
                      changes)
        # Actuarial assumption values from prior sample (specific percentages not from TB)
        for coord in _EBP_ACTUARIAL_CELLS:
            cell = ws[coord]
            if cell.value is not None and str(cell.value).strip() not in {"", "N.A."}:
                cell.value = NA
                changes.append(f"{coord}: sample actuarial assumption cleared")
        # Stale prior-year date column headers  e.g. E36="31st March, 2024" / E53="31st March, 2023"
        for coord in ("E36", "E53", "E70", "E84", "E98", "E107", "E119", "E137", "E154", "E168"):
            cell = ws[coord]
            if cell.value and "2023" in str(cell.value):
                cell.value = NA
                changes.append(f"{coord}: stale 2023 date → N.A.")
        # EBP column header: F13
        _force_set(ws, "F13", "31st March, 2026", changes)

    # ------------------------------------------------------------------ #
    # Other Notes
    # ------------------------------------------------------------------ #
    if sheet_name == "Other Notes":
        # Related party names in Note 36
        for coord in ("B70", "B71", "B72", "B73", "B74"):
            cell = ws[coord]
            if cell.value is not None:
                s = str(cell.value)
                if (
                    "…" in s
                    or "........" in s
                    or "Punit Kumar Sarda" in s
                    or ("XYZ" in s and "Private Limited" in s and "related" not in s.lower())
                ):
                    cell.value = NA
                    changes.append(f"{coord}: related party placeholder/sample name cleared")
        # B29 in Note 33 (Disclosure 186(4)) — "XYZ Limited" sample controlled entity
        cell = ws["B29"]
        if cell.value and "XYZ Limited" in str(cell.value) and "Private" not in str(cell.value):
            cell.value = NA
            changes.append("B29: sample controlled entity 'XYZ Limited' cleared")
        # Note 38 (CSR) — unsupported assertion with specific year
        _replace_text(ws, "B112",
                      "financial year ending March 31, 2025",
                      "financial year ending March 31, 2026",
                      changes)
        # Note 39 management assertion — acceptable boilerplate, retain as-is
        # Note 44(vii) — replace assertion with TB-neutral version
        cell = ws["B155"]
        if cell.value and "financial year ending March 31, 2025" in str(cell.value):
            cell.value = str(cell.value).replace(
                "financial year ending March 31, 2025",
                "financial year ending March 31, 2026",
            )
            changes.append("B155: year corrected")
        # Ratio note: column headers
        for coord, repl in [
            ("F33", "For the year ended March 31, 2026"),
            ("E8",  "31st March, 2026"),
            ("F8",  "31st March, 2025"),
            ("E21", "As at 31st March, 2026"),
        ]:
            cell = ws[coord]
            if cell.value and "2025" in str(cell.value) and "2026" not in str(cell.value):
                cell.value = repl
                changes.append(f"{coord}: date corrected → {repl}")
        # F79 related party transactions column
        _replace_text(ws, "F79",
                      "For the Year ended\n31-Mar-2025",
                      "For the Year ended\n31-Mar-2026",
                      changes)

    # ------------------------------------------------------------------ #
    # Stock Reconciliation
    # ------------------------------------------------------------------ #
    if sheet_name == "Stock Reconciliation ":
        # Period headers
        _replace_text(ws, "E5",
                      "For the year ended 31st March 2025",
                      "For the year ended 31st March 2026",
                      changes)
        _replace_text(ws, "I5",
                      "For the year ended 31st March 2024",
                      "For the year ended 31st March 2025",
                      changes)
        # Note 47 SMC statement — year reference
        cell = ws["B22"]
        if cell.value:
            updated = str(cell.value).replace("March 31, 2025", "March 31, 2026")
            if updated != str(cell.value):
                cell.value = updated
                changes.append("B22: year corrected in SMC note")
        # B23 stray backslash (already handled by generic loop but confirm)
        if ws["B23"].value in ("\\", "\\'", "'"):
            ws["B23"].value = _BLANK
            changes.append("B23: stray backslash cleared")
        # Quarter date headers — raw datetimes already replaced by generic loop
        # But also fix the text Q-period labels if any remain
        for coord in ("E9", "E12", "E15", "E18", "I9", "I12", "I15", "I18"):
            cell = ws[coord]
            if cell.value == NA:
                # Already replaced by datetime handler — good
                pass

    # ------------------------------------------------------------------ #
    # Borrowing
    # ------------------------------------------------------------------ #
    if sheet_name == "Borrowing":
        # Column headers
        for coord, repl in [
            ("E5", "As at 31st March 2026"),
            ("G5", "As at 31st March 2025"),
        ]:
            cell = ws[coord]
            if cell.value and "2025" in str(cell.value) and "2026" not in str(cell.value):
                cell.value = repl
                changes.append(f"{coord}: date corrected → {repl}")

    # ------------------------------------------------------------------ #
    # EPS
    # ------------------------------------------------------------------ #
    if sheet_name == "EPS":
        _replace_text(ws, "C5",
                      "For the period ended at March 31, 2025",
                      "For the period ended at March 31, 2026",
                      changes)
        # EPS values — not available from TB (numeric cells B9-B16 are blank by default;
        # mark N.A. label rows)
        for coord in ("B9", "B10", "B11", "B12", "B13", "B14", "B15", "B16"):
            # Only add N.A. if cell is empty (don't overwrite actual numbers)
            pass  # label rows — keep structural labels as-is

    # ------------------------------------------------------------------ #
    # Ratio_working (internal working sheet)
    # ------------------------------------------------------------------ #
    if sheet_name == "Ratio_working":
        # Column headers FY/PY: rename to be period-specific
        _force_set(ws, "C1", "2025-26", changes)
        _force_set(ws, "D1", "2024-25 (N.A.)", changes)


# --------------------------------------------------------------------------- #
# Helpers
# --------------------------------------------------------------------------- #

def _replace_text(ws: "Worksheet", coord: str, old: str, new: str, changes: list[str]) -> None:
    cell = ws[coord]
    if cell.value and old in str(cell.value):
        cell.value = str(cell.value).replace(old, new)
        changes.append(f"{coord}: '{old[:50]}' → '{new[:50]}'")


def _clear_sample_text(ws: "Worksheet", coord: str, expected: str, replacement: str, changes: list[str]) -> None:
    cell = ws[coord]
    if cell.value is None:
        return
    s = str(cell.value)
    if s == expected or (expected[:30] in s):
        cell.value = replacement
        changes.append(f"{coord}: sample text cleared → '{replacement[:60]}'")


def _force_set(ws: "Worksheet", coord: str, value: str, changes: list[str]) -> None:
    cell = ws[coord]
    if cell.value != value:
        old = repr(cell.value)[:50]
        cell.value = value
        changes.append(f"{coord}: {old} → '{value}'")
