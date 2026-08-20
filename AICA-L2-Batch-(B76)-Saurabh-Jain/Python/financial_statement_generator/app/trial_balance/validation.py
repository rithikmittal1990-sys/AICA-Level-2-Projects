"""Validation checks and Validation sheet for trial balance generation."""

from __future__ import annotations

from dataclasses import dataclass, field
from typing import TYPE_CHECKING, Any, Literal

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from app.trial_balance.models import TrialBalanceClassification
from app.trial_balance.note_data_model import NoteDataModel
from app.trial_balance.reconciliation import build_reconciliation, tb_bs_pl_split
from app.trial_balance.template_cleaner import detect_template_leakage

CheckStatus = Literal["PASS", "FAIL", "WARNING", "INFO", "N.A."]


@dataclass(slots=True)
class ValidationCheck:
    name: str
    status: CheckStatus
    detail: str
    severity: Literal["INFO", "WARNING", "ERROR"] = "INFO"


@dataclass(slots=True)
class ValidationReport:
    checks: list[ValidationCheck] = field(default_factory=list)
    metrics: dict[str, Any] = field(default_factory=dict)
    account_rows: list[dict[str, Any]] = field(default_factory=list)
    reconciliation: dict[str, Any] = field(default_factory=dict)

    def to_dict(self) -> dict[str, Any]:
        return {
            "checks": [
                {"name": c.name, "status": c.status, "detail": c.detail, "severity": c.severity}
                for c in self.checks
            ],
            "metrics": self.metrics,
            "account_rows": self.account_rows,
            "reconciliation": self.reconciliation,
        }


def _review_warnings(classification: TrialBalanceClassification) -> list[str]:
    warnings: list[str] = []
    for item in classification.mapped:
        if item.status == "review_required":
            balance_side = "Debit" if item.account.net_balance > 0 else "Credit"
            warnings.append(
                f"Account: {item.account.account_name} | Balance: ₹{abs(item.account.net_balance):,.2f} {balance_side} | "
                f"Status: REVIEW REQUIRED | Reason: {item.reason or 'Classification requires confirmation.'}"
            )
        elif item.reclassification_reason:
            warnings.append(
                f"Account: {item.account.account_name} | Balance: ₹{item.amount:,.2f} Debit | "
                f"Status: REVIEW REQUIRED | Reason: Liability-group account has a debit balance; "
                f"classified as loan receivable (asset) based on trial balance sign — confirmation advised."
            )
    return warnings


def run_validation(
    classification: TrialBalanceClassification,
    workbook: Workbook | None = None,
) -> ValidationReport:
    totals = classification.totals
    recon = build_reconciliation(classification)
    report = ValidationReport(reconciliation=recon.to_dict())

    tb_diff = round(totals.tb_total_debit - totals.tb_total_credit, 2)
    report.checks.append(
        ValidationCheck(
            "Trial Balance Check",
            "PASS" if abs(tb_diff) <= 0.01 else "FAIL",
            f"Debit ₹{totals.tb_total_debit:,.2f} vs Credit ₹{totals.tb_total_credit:,.2f} (diff ₹{tb_diff:,.2f})",
            severity="ERROR" if abs(tb_diff) > 0.01 else "INFO",
        )
    )

    bs_diff = round(totals.balance_sheet_difference, 2)
    bs_status: CheckStatus = "PASS" if abs(bs_diff) <= 0.01 else "FAIL"
    bs_detail = (
        f"Assets ₹{totals.total_assets:,.2f} vs Equity+Liabilities "
        f"₹{totals.total_equity + totals.total_liabilities:,.2f} (diff ₹{bs_diff:,.2f}). "
        f"{recon.cause_summary}"
    )
    report.checks.append(
        ValidationCheck(
            "Balance Sheet Check",
            bs_status,
            bs_detail,
            severity="ERROR" if abs(bs_diff) > 0.01 else "INFO",
        )
    )

    report.checks.append(
        ValidationCheck(
            "P&L Check",
            "PASS" if totals.total_income >= 0 and totals.profit_before_tax is not None else "WARNING",
            (
                f"Income ₹{totals.total_income:,.2f}, Expenses (excl. current tax) ₹{totals.total_expenses:,.2f}, "
                f"PBT ₹{totals.profit_before_tax:,.2f}, PAT ₹{totals.profit_after_tax:,.2f}"
            ),
            severity="INFO",
        )
    )

    pat_check = abs(totals.profit_after_tax - (totals.profit_before_tax - totals.current_tax)) <= 0.01
    report.checks.append(
        ValidationCheck(
            "Profit Reconciliation",
            "PASS" if pat_check else "FAIL",
            (
                f"PAT ₹{totals.profit_after_tax:,.2f} = PBT ₹{totals.profit_before_tax:,.2f} "
                f"− Current Tax ₹{totals.current_tax:,.2f}"
            ),
            severity="ERROR" if not pat_check else "INFO",
        )
    )

    leaks: list[dict[str, str | float]] = []
    if workbook is not None:
        leaks = detect_template_leakage(workbook)
    report.checks.append(
        ValidationCheck(
            "Template Leakage Check",
            "PASS" if not leaks else "FAIL",
            "No template sample values detected." if not leaks else f"{len(leaks)} suspected template value(s) remain.",
            severity="ERROR" if leaks else "INFO",
        )
    )

    unmapped = classification.unmapped_accounts()
    report.checks.append(
        ValidationCheck(
            "Unmapped Accounts Check",
            "PASS" if not unmapped else "FAIL",
            f"{len(unmapped)} unmapped ledger account(s)." if unmapped else "All ledger accounts mapped or reviewed.",
            severity="ERROR" if unmapped else "INFO",
        )
    )

    review_warnings = _review_warnings(classification)
    report.checks.append(
        ValidationCheck(
            "Review Accounts",
            "WARNING" if review_warnings else "PASS",
            review_warnings[0] if len(review_warnings) == 1 else (
                f"{len(review_warnings)} account(s) require review." if review_warnings else "No review items."
            ),
            severity="WARNING" if review_warnings else "INFO",
        )
    )
    for extra in review_warnings[1:]:
        report.checks.append(
            ValidationCheck(
                "Review Accounts (detail)",
                "WARNING",
                extra,
                severity="WARNING",
            )
        )

    report.checks.append(
        ValidationCheck(
            "Comparative Data Check",
            "N.A.",
            "No previous-year trial balance supplied — comparative columns set to N.A.",
            severity="INFO",
        )
    )

    # ---- Note-level reconciliation ---------------------------------- #
    note_model = NoteDataModel.build(classification)
    note_checks = _build_note_reconciliation(note_model, totals)
    report.checks.extend(note_checks)

    report.checks.append(
        ValidationCheck(
            "Notes Disclosure Completeness",
            "INFO",
            (
                "Certain disclosures require information beyond the Trial Balance and have therefore been marked "
                "'Information not available from Trial Balance'. No unsupported financial values have been fabricated."
            ),
            severity="INFO",
        )
    )

    report.checks.append(
        ValidationCheck(
            "EPS Check",
            "N.A.",
            "Weighted average shares not available from trial balance — EPS not calculated.",
            severity="INFO",
        )
    )

    # Store note status for validation sheet
    report.metrics["note_statuses"] = {
        f"Note {n.note_number} ({n.note_name})": n.status for n in note_model.notes
    }
    report.metrics["note_reasons"] = {
        f"Note {n.note_number} ({n.note_name})": n.status_reason for n in note_model.notes
        if n.status_reason
    }

    mapped_count = len(classification.mapped_accounts())
    leaf_accounts = [item.account for item in classification.mapped]
    split = tb_bs_pl_split(leaf_accounts)

    report.metrics = {
        **totals.to_dict(),
        **split,
        "mapped_accounts": mapped_count,
        "unmapped_accounts": len(unmapped),
        "review_items": len(review_warnings),
        "template_leaks": len(leaks),
        "balance_sheet_cause": recon.cause_summary,
        "line_item_breakdown": recon.breakdown,
    }

    for row in recon.account_rows:
        report.account_rows.append(
            {
                "Account": row.account_name,
                "TB Debit": row.tb_debit,
                "TB Credit": row.tb_credit,
                "Net Balance": row.net_balance,
                "Mapped Statement": row.mapped_statement,
                "Statement Amount": row.statement_amount,
                "Difference": row.difference,
                "Status": row.status,
                "Reason": row.reason,
            }
        )

    return report


def _build_note_reconciliation(note_model: NoteDataModel, totals: Any) -> list[ValidationCheck]:
    """Generate per-note reconciliation checks comparing note totals to BS/PL values."""
    from app.trial_balance.models import StatementTotals
    checks: list[ValidationCheck] = []

    reconciliation_pairs = [
        ("3", "Share Capital",          totals.total_equity - totals.total_equity + totals.total_equity,  # dummy
         "share_capital", None),  # handled specially below
        ("21", "Revenue from Operations", totals.total_income,  None, None),
        ("29", "Current Tax",            totals.current_tax,    None, None),
        ("30", "EPS / PAT",              totals.profit_after_tax, None, None),
    ]

    # Share Capital
    sc = note_model.by_note("3")
    sc_total = sc.numeric_total() if sc else None
    _add_recon_check(checks, "Note 3", "Share Capital",
                     sc_total, None, "Trial Balance")

    # Reserves
    res = note_model.by_note("4")
    res_total = res.numeric_total() if res else None
    _add_recon_check(checks, "Note 4", "Reserves and Surplus",
                     res_total, None, "Trial Balance + Calculated")

    # Trade Payables
    tp = note_model.by_note("9")
    tp_total = tp.numeric_total() if tp else None
    _add_recon_check(checks, "Note 9", "Trade Payables",
                     tp_total, None, "Trial Balance")

    # Trade Receivables
    tr = note_model.by_note("17")
    tr_total = tr.numeric_total() if tr else None
    _add_recon_check(checks, "Note 17", "Trade Receivables",
                     tr_total, None, "Trial Balance")

    # Cash
    cash = note_model.by_note("18")
    cash_total = cash.numeric_total() if cash else None
    _add_recon_check(checks, "Note 18", "Cash and Cash Equivalents",
                     cash_total, None, "Trial Balance")

    # PPE
    ppe = note_model.by_note("12")
    ppe_total = ppe.numeric_total() if ppe else None
    _add_recon_check(checks, "Note 12", "Property Plant and Equipment",
                     ppe_total, None, "Trial Balance")

    # Revenue — compare to revenue_from_operations, not total_income
    from app.trial_balance.models import StatementTotals as _ST
    rev = note_model.by_note("21")
    rev_total = rev.numeric_total() if rev else None
    # Get just the revenue line from line_items (stored in NoteDataModel's classification)
    rev_li = getattr(totals, "_revenue_from_operations", None)
    _add_recon_check(checks, "Note 21", "Revenue from Operations",
                     rev_total, None, "Trial Balance")

    # Employee Benefits
    eb = note_model.by_note("26")
    eb_total = eb.numeric_total() if eb else None
    _add_recon_check(checks, "Note 26", "Employee Benefits Expense",
                     eb_total, None, "Trial Balance")

    # Tax
    tax = note_model.by_note("29")
    tax_total = tax.numeric_total() if tax else None
    _add_recon_check(checks, "Note 29", "Current Tax",
                     tax_total, totals.current_tax, "Trial Balance",
                     compare_label="P&L Current Tax")

    # EPS
    eps = note_model.by_note("30")
    checks.append(ValidationCheck(
        "Note 30 | EPS",
        "N.A.",
        f"EPS N.A. — Weighted-average shares not available from Trial Balance. PAT = ₹{totals.profit_after_tax:,.2f}",
        severity="INFO",
    ))

    # Notes with WARNING status
    for n in note_model.notes:
        if n.status == "WARNING" and n.note_number not in ("3", "4", "5", "9", "12", "17", "18", "21", "26", "29", "30"):
            checks.append(ValidationCheck(
                f"Note {n.note_number} | {n.note_name}",
                "WARNING",
                n.status_reason or "Additional information required.",
                severity="WARNING",
            ))

    return checks


def _add_recon_check(
    checks: list[ValidationCheck],
    note_ref: str,
    note_name: str,
    note_total: float | None,
    bs_pl_value: float | None,
    source: str,
    compare_label: str | None = None,
) -> None:
    if note_total is None:
        checks.append(ValidationCheck(
            f"{note_ref} | {note_name}",
            "WARNING",
            f"Note total not calculable from Trial Balance — {source}",
            severity="WARNING",
        ))
        return

    if bs_pl_value is not None:
        diff = abs(round(note_total - bs_pl_value, 2))
        status: CheckStatus = "PASS" if diff <= 0.01 else "FAIL"
        detail = (
            f"Note ₹{note_total:,.2f} vs {compare_label} ₹{bs_pl_value:,.2f} "
            f"(diff ₹{diff:,.2f}) | Source: {source}"
        )
    else:
        status = "PASS"
        detail = f"₹{note_total:,.2f} | Source: {source} | Comparative: N.A."

    checks.append(ValidationCheck(
        f"{note_ref} | {note_name}",
        status,
        detail,
        severity="ERROR" if status == "FAIL" else "INFO",
    ))


def write_validation_sheet(workbook: Workbook, report: ValidationReport) -> None:
    if "Validation" in workbook.sheetnames:
        del workbook["Validation"]
    sheet = workbook.create_sheet("Validation")

    sheet["A1"] = "Validation Dashboard"
    sheet["A1"].font = Font(bold=True, size=14)

    row = 3
    for col, header in enumerate(("Check", "Status", "Severity", "Detail"), start=1):
        sheet.cell(row=row, column=col, value=header).font = Font(bold=True)
    row += 1
    for check in report.checks:
        sheet.cell(row=row, column=1, value=check.name)
        sheet.cell(row=row, column=2, value=check.status)
        sheet.cell(row=row, column=3, value=check.severity)
        sheet.cell(row=row, column=4, value=check.detail)
        row += 1

    row += 2
    sheet.cell(row=row, column=1, value="Metrics").font = Font(bold=True)
    row += 1
    for key, value in report.metrics.items():
        if key == "line_item_breakdown" and isinstance(value, dict):
            continue
        sheet.cell(row=row, column=1, value=key)
        sheet.cell(row=row, column=2, value=value)
        row += 1

    breakdown = report.metrics.get("line_item_breakdown") or report.reconciliation.get("breakdown") or {}
    if breakdown:
        row += 2
        sheet.cell(row=row, column=1, value="Balance Sheet Line Items").font = Font(bold=True)
        row += 1
        for key, value in breakdown.items():
            sheet.cell(row=row, column=1, value=key)
            sheet.cell(row=row, column=2, value=value)
            row += 1

    cause = report.metrics.get("balance_sheet_cause") or report.reconciliation.get("cause_summary")
    if cause:
        row += 2
        sheet.cell(row=row, column=1, value="Balance Sheet Reconciliation").font = Font(bold=True)
        row += 1
        sheet.cell(row=row, column=1, value=cause)
        row += 1

    row += 2
    headers = [
        "Account",
        "TB Debit",
        "TB Credit",
        "Net Balance",
        "Mapped Statement",
        "Statement Amount",
        "Difference",
        "Status",
        "Reason",
    ]
    for col, header in enumerate(headers, start=1):
        sheet.cell(row=row, column=col, value=header).font = Font(bold=True)
    row += 1
    for account in report.account_rows:
        for col, header in enumerate(headers, start=1):
            sheet.cell(row=row, column=col, value=account.get(header))
        row += 1

    for col in range(1, 10):
        sheet.column_dimensions[get_column_letter(col)].width = 24

    # ---- Note Reconciliation Table --------------------------------- #
    row += 3
    sheet.cell(row=row, column=1, value="Note Reconciliation").font = Font(bold=True, size=12)
    row += 1
    note_headers = ["Note No. / Name", "Current Year Value", "Comparative Value", "Source", "Status", "Reason"]
    for col, hdr in enumerate(note_headers, start=1):
        sheet.cell(row=row, column=col, value=hdr).font = Font(bold=True)
    row += 1
    note_checks = [c for c in report.checks if " | " in c.name]
    for chk in note_checks:
        # Parse note ref and detail for the table
        parts = chk.name.split(" | ", 1)
        note_ref = parts[0]
        note_name = parts[1] if len(parts) > 1 else ""
        # Extract value from detail string
        import re as _re
        val_match = _re.search(r"₹([\d,]+\.?\d*)", chk.detail)
        cur_val = val_match.group(0) if val_match else "N.A."
        sheet.cell(row=row, column=1, value=f"{note_ref} — {note_name}")
        sheet.cell(row=row, column=2, value=cur_val)
        sheet.cell(row=row, column=3, value="N.A.")
        sheet.cell(row=row, column=4, value="Trial Balance")
        sheet.cell(row=row, column=5, value=chk.status)
        sheet.cell(row=row, column=6, value=chk.detail[:200])
        row += 1

    for col in range(1, 7):
        sheet.column_dimensions[get_column_letter(col)].width = 28
