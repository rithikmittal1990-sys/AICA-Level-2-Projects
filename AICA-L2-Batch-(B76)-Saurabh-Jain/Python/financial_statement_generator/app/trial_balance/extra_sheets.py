"""
Extra Sheets
============
Creates two informational sheets in the generated workbook:

1. TB Mapping   — every TB account → Statement Head, Note, Confidence, Status
2. Disclosure Status — per-disclosure table: can derive from TB?, value, status
"""

from __future__ import annotations

from typing import TYPE_CHECKING

from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from app.trial_balance.note_data_model import NoteDataModel

if TYPE_CHECKING:
    from openpyxl import Workbook
    from openpyxl.worksheet.worksheet import Worksheet

_PASS_FILL    = PatternFill("solid", fgColor="C6EFCE")
_WARN_FILL    = PatternFill("solid", fgColor="FFEB9C")
_FAIL_FILL    = PatternFill("solid", fgColor="FFC7CE")
_NA_FILL      = PatternFill("solid", fgColor="DDDDDD")
_HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT  = Font(bold=True, color="FFFFFF", size=10)
_BOLD         = Font(bold=True)


def _status_fill(status: str) -> PatternFill:
    s = status.upper()
    if s in ("PASS",):                return _PASS_FILL
    if s in ("WARNING", "REVIEW"):    return _WARN_FILL
    if s in ("FAIL",):                return _FAIL_FILL
    return _NA_FILL


def _header_row(ws: "Worksheet", row: int, headers: list[str]) -> None:
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font  = _HEADER_FONT
        c.fill  = _HEADER_FILL
        c.alignment = Alignment(horizontal="center", wrap_text=True)


def _autowidth(ws: "Worksheet", widths: list[int]) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# --------------------------------------------------------------------------- #
# Sheet 1: TB Mapping
# --------------------------------------------------------------------------- #

def write_tb_mapping_sheet(workbook: "Workbook", model: NoteDataModel) -> None:
    if "TB Mapping" in workbook.sheetnames:
        del workbook["TB Mapping"]
    ws = workbook.create_sheet("TB Mapping")

    ws["A1"] = f"Trial Balance Account Mapping — {model.company_name}"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = f"Period: {model.period_label}   |   Reporting Date: {model.reporting_date}"
    ws["A2"].font = Font(italic=True, color="444444")

    headers = [
        "Account Name", "Parent Group",
        "TB Debit (₹)", "TB Credit (₹)", "Net Balance (₹)", "Balance Side",
        "Statement Section", "Statement Head", "Note No.", "Aggregate Key",
        "Source", "Confidence", "Status", "Notes / Review Reason",
    ]
    _header_row(ws, 4, headers)

    row = 5
    for m in model.all_mapped:
        a   = m.account
        net = a.net_balance
        side = "Debit" if net > 0 else ("Credit" if net < 0 else "Zero")

        # Confidence
        if m.reclassification_reason:
            confidence = "Medium"
        elif m.status == "review_required":
            confidence = "Low"
        elif m.status == "unmapped":
            confidence = "Unmapped"
        else:
            confidence = "High"

        # Status label
        if m.status == "mapped":
            status_lbl = "PASS" if not m.reclassification_reason else "REVIEW"
        elif m.status == "review_required":
            status_lbl = "REVIEW"
        else:
            status_lbl = "UNMAPPED"

        reason = (m.reclassification_reason or m.reason or "")[:150]

        values = [
            a.account_name,
            a.parent_group or "",
            round(a.debit, 2),
            round(a.credit, 2),
            round(abs(net), 2),
            side,
            m.statement,
            m.statement_head,
            m.note or "",
            m.note or "",
            "Trial Balance",
            confidence,
            status_lbl,
            reason,
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            if col == 13:
                cell.fill = _status_fill(status_lbl)
        row += 1

    # Totals row
    ws.cell(row=row, column=1, value="TOTALS").font = _BOLD
    for col_idx, col_letter in [(3, "C"), (4, "D"), (5, "E")]:
        ws.cell(row=row, column=col_idx,
                value=f"=SUM({col_letter}5:{col_letter}{row-1})")

    _autowidth(ws, [32, 20, 15, 15, 15, 10, 18, 28, 10, 28, 14, 12, 10, 50])
    ws.freeze_panes = "A5"


# --------------------------------------------------------------------------- #
# Sheet 2: Disclosure Status
# --------------------------------------------------------------------------- #

_DISCLOSURES = [
    # (Disclosure, Can derive from TB?, Note ref, additional_info_required)
    ("Share Capital",                  True,  "3",    "No"),
    ("Number of Shares",               True,  "3",    "Face value assumed ₹10"),
    ("Authorised Capital",             False, "3",    "Yes — not in Trial Balance"),
    ("Shares held by shareholders",    False, "3",    "Yes — shareholder register required"),
    ("Reserves and Surplus",           True,  "4",    "No"),
    ("Opening Reserves movement",      False, "4",    "Yes — prior year TB required"),
    ("Borrowings",                     False, "5",    "Yes — Ankit Loan review required"),
    ("Borrowing terms / security",     False, "5",    "Yes — loan agreement required"),
    ("Deferred Tax Asset",             True,  "6",    "Partial — timing diff breakup required"),
    ("Deferred Tax Income",            True,  "6",    "Partial — timing diff breakup required"),
    ("Trade Payables",                 True,  "9",    "No"),
    ("MSME Classification",            False, "9",    "Yes — MSME register required"),
    ("Trade Payable Ageing",           False, "9",    "Yes — invoice-level data required"),
    ("Other Current Liabilities",      True,  "10",   "No"),
    ("Short-term Provisions",          True,  "11",   "No"),
    ("PPE — Net Book Value",           True,  "12",   "No"),
    ("PPE — Gross Block / Movement",   False, "12",   "Yes — fixed asset register required"),
    ("CWIP",                           True,  "13",   "No (Nil from TB)"),
    ("Non-Current Investments",        True,  "14",   "No (Nil from TB)"),
    ("Trade Receivables",              True,  "17",   "No"),
    ("Trade Receivable Ageing",        False, "17",   "Yes — invoice-level data required"),
    ("Cash and Cash Equivalents",      True,  "18",   "No"),
    ("Short-term Loans & Advances",    True,  "19",   "Partial — Ankit Loan review required"),
    ("Other Current Assets",           True,  "20",   "No"),
    ("Revenue from Operations",        True,  "21",   "No"),
    ("Other Income",                   True,  "22",   "No"),
    ("Employee Benefits Expense",      True,  "26",   "No"),
    ("Gratuity / Actuarial details",   False, "26",   "Yes — actuarial report required"),
    ("Depreciation",                   True,  "PPE",  "No"),
    ("Other Expenses",                 True,  "27",   "No"),
    ("Finance Costs",                  True,  "28",   "Partial"),
    ("Current Tax",                    True,  "29",   "No"),
    ("Deferred Tax (P&L)",             True,  "29",   "No"),
    ("EPS",                            False, "30",   "Yes — weighted-average shares required"),
    ("Contingent Liabilities",         False, "35",   "Yes — management confirmation required"),
    ("Related Party Disclosures",      False, "36",   "Yes — related party register required"),
    ("Segment Information",            False, "37",   "Yes — management information required"),
    ("CSR Disclosure",                 False, "38",   "Yes — company information required"),
    ("Events After Reporting Date",    False, "40",   "Yes — management confirmation required"),
    ("Regulatory Declarations",        False, "44",   "Yes — management confirmation required"),
    ("EBP — Gratuity Plan Details",    False, "45",   "Yes — actuarial report required"),
    ("SMC Classification",             False, "47",   "Yes — company information required"),
    ("MSME Quarterly Bank Returns",    False, "46",   "Yes — bank statement required"),
]


def write_disclosure_status_sheet(
    workbook: "Workbook",
    model: NoteDataModel,
    classification_totals: dict,
) -> None:
    if "Disclosure Status" in workbook.sheetnames:
        del workbook["Disclosure Status"]
    ws = workbook.create_sheet("Disclosure Status")

    ws["A1"] = f"Disclosure Status — {model.company_name}"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = f"Period: {model.period_label}   |   Reporting Date: {model.reporting_date}"
    ws["A2"].font = Font(italic=True, color="444444")

    headers = [
        "Disclosure", "Can derive from TB?", "Note No.",
        "Current Year Value", "Comparative Value",
        "Source", "Status", "Additional Information Required",
    ]
    _header_row(ws, 4, headers)

    # Build value map from model notes
    note_value_map: dict[str, str] = {}
    for n in model.notes:
        lbl = f"Note {n.note_number}"
        if n.total is not None:
            note_value_map[n.note_number] = f"₹{n.total:,.2f}"
        elif n.status == "N.A.":
            note_value_map[n.note_number] = "N.A."
        else:
            # Find first numeric line item
            v = next((i.amount for i in n.line_items if isinstance(i.amount, (int, float))), None)
            note_value_map[n.note_number] = f"₹{v:,.2f}" if v is not None else "N.A."

    # Known values for non-note disclosures
    tb_totals = classification_totals
    known_values: dict[str, str] = {
        "Revenue from Operations":   f"₹{tb_totals.get('total_income', 0) - tb_totals.get('deferred_tax', 0) - (tb_totals.get('total_income', 0) - tb_totals.get('total_income', 0)):,.2f}",
        "Depreciation":              f"₹{classification_totals.get('depreciation', 0):,.2f}",
        "Current Tax":               f"₹{tb_totals.get('current_tax', 0):,.2f}",
    }

    row = 5
    for disclosure, can_derive, note_ref, add_info in _DISCLOSURES:
        if can_derive:
            cur_val = note_value_map.get(note_ref, "N.A.")
            source  = "Trial Balance"
            status  = "PASS"
        else:
            cur_val = "N.A."
            source  = "Not available"
            status  = "WARNING"

        ws.cell(row=row, column=1, value=disclosure)
        ws.cell(row=row, column=2, value="Yes" if can_derive else "No")
        ws.cell(row=row, column=3, value=note_ref)
        ws.cell(row=row, column=4, value=cur_val)
        ws.cell(row=row, column=5, value="N.A.")
        ws.cell(row=row, column=6, value=source)
        status_cell = ws.cell(row=row, column=7, value=status)
        status_cell.fill = _status_fill(status)
        ws.cell(row=row, column=8, value=add_info)
        row += 1

    _autowidth(ws, [38, 18, 10, 22, 20, 18, 12, 44])
    ws.freeze_panes = "A5"
