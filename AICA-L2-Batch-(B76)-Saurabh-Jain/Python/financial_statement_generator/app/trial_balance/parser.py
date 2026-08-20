"""Parse hierarchical trial balance workbooks into leaf ledger accounts."""

from __future__ import annotations

import json
import re
import uuid
from dataclasses import dataclass, field
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.config import settings
from app.extraction.exceptions import InvalidTrialBalanceError, TrialBalanceTooLargeError
from app.trial_balance.models import NormalizedAccount

EXCEL_SUFFIXES = {".xlsx", ".xlsm"}
LABEL_RE = re.compile(r"[\t\r\n_]+")
PERIOD_RE = re.compile(
    r"(\d{1,2}[-/]\w{3,9}[-/]\d{2,4})\s*(?:to|-)\s*(\d{1,2}[-/]\w{3,9}[-/]\d{2,4})",
    re.IGNORECASE,
)
# Matches email addresses — these should never be a company name or period
_EMAIL_RE  = re.compile(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}")
# XML carriage-return artifact left by some Excel exporters
_CRLF_JUNK = re.compile(r"_x000D_", re.IGNORECASE)
DEFAULT_GROUP_CONFIG = Path(__file__).resolve().parent.parent / "mapping" / "trial_balance_account_map.json"


@dataclass(slots=True)
class TrialBalanceParseResult:
    company_name: str | None
    period_label: str | None
    sheet_name: str
    stored_path: str
    accounts: list[NormalizedAccount] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    grand_total_debit: float | None = None
    grand_total_credit: float | None = None

    @property
    def leaf_accounts(self) -> list[NormalizedAccount]:
        return [item for item in self.accounts if not item.is_group]


class TrialBalanceParser:
    def __init__(
        self,
        upload_dir: Path | None = None,
        max_upload_bytes: int | None = None,
        group_config: Path | None = None,
    ) -> None:
        self.upload_dir = Path(upload_dir) if upload_dir else settings.upload_dir
        self.max_upload_bytes = max_upload_bytes if max_upload_bytes is not None else settings.max_upload_bytes
        self.upload_dir.mkdir(parents=True, exist_ok=True)
        path = group_config or DEFAULT_GROUP_CONFIG
        payload = json.loads(path.read_text(encoding="utf-8"))
        self.group_labels = {normalize_label(item) for item in payload.get("group_labels") or []}

    def save_upload(self, filename: str, content: bytes) -> Path:
        if not content:
            raise InvalidTrialBalanceError("The uploaded file is empty.")
        if len(content) > self.max_upload_bytes:
            raise TrialBalanceTooLargeError(
                f"Trial balance exceeds the maximum upload size of {self.max_upload_bytes} bytes."
            )
        suffix = Path(filename or "trial_balance.xlsx").suffix.lower()
        if suffix not in EXCEL_SUFFIXES:
            raise InvalidTrialBalanceError("Upload a .xlsx trial balance workbook.")
        safe_name = Path(filename or "trial_balance.xlsx").name
        stored = self.upload_dir / f"{uuid.uuid4().hex}_{safe_name}"
        stored.write_bytes(content)
        return stored

    def read_upload(self, filename: str, content: bytes) -> TrialBalanceParseResult:
        stored = self.save_upload(filename, content)
        return self.read_path(stored)

    def read_path(self, path: Path) -> TrialBalanceParseResult:
        try:
            workbook = load_workbook(path, data_only=True, read_only=True)
        except Exception as exc:
            raise InvalidTrialBalanceError("The trial balance workbook could not be opened.") from exc
        try:
            if not workbook.sheetnames:
                raise InvalidTrialBalanceError("The trial balance workbook has no sheets.")
            sheet = workbook[workbook.sheetnames[0]]
            company_name = _first_text(sheet, ("A1", "A2", "B4"))
            period_label = _first_text(sheet, ("A3", "B5"))
            header_row = _find_header_row(sheet)
            if header_row is None:
                raise InvalidTrialBalanceError(
                    "Could not find a trial balance header row with Particulars and Debit/Credit columns."
                )
            accounts = _parse_accounts(sheet, header_row, self.group_labels)
            if not accounts:
                raise InvalidTrialBalanceError("No account rows were found in the trial balance.")
            gt_debit, gt_credit = _read_grand_total(sheet)
            warnings: list[str] = []
            if company_name is None:
                warnings.append("Company name could not be read from the trial balance.")
            if period_label is None:
                warnings.append("Reporting period could not be read from the trial balance.")
            return TrialBalanceParseResult(
                company_name=company_name,
                period_label=period_label,
                sheet_name=sheet.title,
                stored_path=str(path),
                accounts=accounts,
                warnings=warnings,
                grand_total_debit=gt_debit,
                grand_total_credit=gt_credit,
            )
        finally:
            workbook.close()


def normalize_label(value: Any) -> str:
    text = "" if value is None else str(value)
    text = LABEL_RE.sub(" ", text)
    return " ".join(text.split()).strip().lower()


def normalize_account_label(value: Any) -> str:
    text = "" if value is None else str(value)
    text = _CRLF_JUNK.sub("", text)    # strip XML carriage-return artifacts
    text = LABEL_RE.sub(" ", text)
    return " ".join(text.split()).strip()


def _read_grand_total(sheet) -> tuple[float | None, float | None]:
    max_row = sheet.max_row or 1
    for row in range(max_row, max(1, max_row - 10), -1):
        label = normalize_label(sheet.cell(row=row, column=1).value)
        if label.startswith("grand total") or label == "total":
            return _as_amount(sheet.cell(row=row, column=2).value), _as_amount(sheet.cell(row=row, column=3).value)
    return None, None


def _parse_accounts(sheet, header_row: int, group_labels: set[str]) -> list[NormalizedAccount]:
    accounts: list[NormalizedAccount] = []
    current_group: str | None = None
    max_row = sheet.max_row or header_row
    for row in range(header_row + 1, max_row + 1):
        label = normalize_account_label(sheet.cell(row=row, column=1).value)
        if not label:
            continue
        label_key = normalize_label(label)
        if label_key.startswith("grand total") or label_key == "total":
            continue
        debit = _as_amount(sheet.cell(row=row, column=2).value) or 0.0
        credit = _as_amount(sheet.cell(row=row, column=3).value) or 0.0
        if debit == 0 and credit == 0:
            continue
        is_group = label_key in group_labels
        if is_group:
            current_group = label
            accounts.append(
                NormalizedAccount(
                    account_name=label,
                    parent_group=current_group,
                    debit=debit,
                    credit=credit,
                    net_balance=round(debit - credit, 2),
                    balance_type=_balance_type(debit, credit),
                    row=row,
                    is_group=True,
                )
            )
            continue
        accounts.append(
            NormalizedAccount(
                account_name=label,
                parent_group=current_group,
                debit=debit,
                credit=credit,
                net_balance=round(debit - credit, 2),
                balance_type=_balance_type(debit, credit),
                row=row,
                is_group=False,
            )
        )
    return accounts


def _balance_type(debit: float, credit: float) -> str:
    net = debit - credit
    if abs(net) < 0.005:
        return "zero"
    return "debit" if net > 0 else "credit"


def _clean_cell_text(raw: Any) -> str:
    """Strip junk (tabs, newlines, _x000D_ XML artifacts) and return clean text."""
    if raw is None:
        return ""
    text = str(raw)
    text = _CRLF_JUNK.sub("", text)       # remove _x000D_ carriage-return artifacts
    text = LABEL_RE.sub(" ", text)         # collapse tabs / newlines
    return " ".join(text.split()).strip()


_SKIP_VALUES = {"trial balance", "particulars", "closing balance", "debit", "credit"}


def _first_text(sheet, coordinates: tuple[str, ...]) -> str | None:
    for coordinate in coordinates:
        raw   = sheet[coordinate].value
        value = _clean_cell_text(raw)
        if not value:
            continue
        if value.lower() in _SKIP_VALUES:
            continue
        # Reject cells that contain an email address — they are leftover from a
        # previous document that happened to use the same cell position.
        if _EMAIL_RE.search(value):
            continue
        if PERIOD_RE.search(value):
            return value
        if len(value) >= 3:
            return value
    return None


def _find_header_row(sheet) -> int | None:
    for row in range(1, min(sheet.max_row or 1, 30) + 1):
        particulars = normalize_label(sheet.cell(row=row, column=1).value)
        if particulars != "particulars":
            continue
        headers = {
            normalize_label(sheet.cell(row=row, column=col).value)
            for col in range(1, min(sheet.max_column or 1, 8) + 1)
        }
        if "debit" in headers and "credit" in headers:
            return row + 1 if "closing balance" in headers else row
        for probe in range(row, row + 3):
            subheaders = {
                normalize_label(sheet.cell(row=probe, column=col).value)
                for col in range(1, min(sheet.max_column or 1, 8) + 1)
            }
            if "debit" in subheaders and "credit" in subheaders:
                return probe
    return None


def _as_amount(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace(",", "").strip()
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def is_trial_balance_upload(filename: str) -> bool:
    return Path(filename or "").suffix.lower() in EXCEL_SUFFIXES


def closing_amount(debit: float | None, credit: float | None, *, nature: str) -> float | None:
    d = float(debit or 0)
    c = float(credit or 0)
    if d == 0 and c == 0:
        return None
    if nature in {"asset", "expense"}:
        return d - c
    return c - d


def financial_year_from_period(period_label: str | None) -> str | None:
    if not period_label:
        return None
    match = PERIOD_RE.search(period_label)
    if not match:
        return period_label
    end = match.group(2)
    # 31-Mar-26 -> FY 2025-26
    parts = re.split(r"[-/]", end)
    if len(parts) >= 3 and len(parts[-1]) == 2:
        year = int(parts[-1])
        year += 2000 if year < 100 else 0
        return f"FY {year - 1}-{str(year)[-2:]}"
    return period_label
