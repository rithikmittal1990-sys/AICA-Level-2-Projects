"""Generate financial statements from trial balance only."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.config import settings
from app.excel.company_branding import apply_company_branding, template_company_names
from app.excel.formatting import write_text, write_value
from app.excel.template_manager import TemplateManager, file_sha256
from app.excel.trial_balance_workbook import remove_excluded_sheets, sanitize_trial_balance_workbook
from app.trial_balance.classifier import TrialBalanceAccountClassifier
from app.trial_balance.parser import TrialBalanceParser
from app.trial_balance.statement_builder import build_cell_writes, writes_to_placements
from app.trial_balance.extra_sheets import write_disclosure_status_sheet, write_tb_mapping_sheet
from app.trial_balance.note_data_model import NoteDataModel
from app.trial_balance.note_writer import write_notes_from_model
from app.trial_balance.notes_auditor import audit_notes_sheets
from app.trial_balance.template_cleaner import clear_comparative_columns, clear_template_financial_data
from app.trial_balance.validation import run_validation, write_validation_sheet

DEFAULT_OUTPUT = "Financial_Statements_Corrected.xlsx"
EXCLUDED_SHEETS = ("Cash Flow",)


@dataclass(slots=True)
class GenerationOutput:
    path: Path
    classification: Any
    validation: Any
    written_count: int


class TrialBalanceStatementGenerator:
    """Copy template layout, clear sample data, write only trial balance values."""

    def __init__(
        self,
        output_dir: Path | None = None,
        template_path: Path | None = None,
    ) -> None:
        self.output_dir = Path(output_dir) if output_dir else settings.output_dir
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.template_manager = TemplateManager(template_path)
        self.parser = TrialBalanceParser()
        self.classifier = TrialBalanceAccountClassifier()

    def generate_from_path(
        self,
        trial_balance_path: str | Path,
        *,
        output_filename: str = DEFAULT_OUTPUT,
    ) -> GenerationOutput:
        parsed = self.parser.read_path(Path(trial_balance_path))
        classification = self.classifier.classify(parsed)
        return self.generate(classification, output_filename=output_filename)

    def generate(
        self,
        classification: Any,
        *,
        output_filename: str = DEFAULT_OUTPUT,
    ) -> GenerationOutput:
        original_hash = file_sha256(self.template_manager.template_path)
        destination = self.output_dir / output_filename
        copied = self.template_manager.copy_template(destination)

        writes = build_cell_writes(classification)
        workbook = load_workbook(copied, data_only=False, rich_text=True)
        written_count = 0
        try:
            remove_excluded_sheets(workbook, EXCLUDED_SHEETS)
            clear_template_financial_data(workbook)
            clear_comparative_columns(workbook)

            company = classification.company_name or ""
            if company:
                apply_company_branding(
                    workbook,
                    company,
                    template_names=template_company_names(),
                    full_workbook=True,
                )
                sanitize_trial_balance_workbook(
                    workbook,
                    company_name=company,
                    period_label=classification.period_label,
                    template_names=template_company_names(),
                )

            for write in writes:
                sheet_name = _resolve_sheet(workbook, write.sheet)
                if sheet_name is None:
                    continue
                sheet = workbook[sheet_name]
                if isinstance(write.value, str):
                    result = write_text(sheet, write.cell, write.value, overwrite_formula=True)
                elif write.value is None:
                    continue
                else:
                    result = write_value(sheet, write.cell, write.value, overwrite_formula=True)
                if result.written:
                    written_count += 1

            # Build central note data model and render into all Notes sheets
            note_model = NoteDataModel.build(classification)
            write_notes_from_model(workbook, note_model)

            # Textual / template audit — second-pass cleanup of dates, sample text, wrong labels
            audit_notes_sheets(
                workbook,
                company_name=classification.company_name,
                period_label=classification.period_label,
            )

            # Build extra auditable sheets
            write_tb_mapping_sheet(workbook, note_model)
            write_disclosure_status_sheet(
                workbook, note_model,
                {**classification.line_items, **classification.totals.to_dict()},
            )

            validation = run_validation(classification, workbook)
            write_validation_sheet(workbook, validation)
            workbook.save(copied)
        finally:
            workbook.close()

        if file_sha256(self.template_manager.template_path) != original_hash:
            raise RuntimeError("Template was modified during generation.")

        # Re-open for leakage check after save
        workbook = load_workbook(copied, data_only=True)
        try:
            validation = run_validation(classification, workbook)
        finally:
            workbook.close()

        return GenerationOutput(
            path=copied,
            classification=classification,
            validation=validation,
            written_count=written_count,
        )


def _resolve_sheet(workbook, preferred: str) -> str | None:
    if preferred in workbook.sheetnames:
        return preferred
    wanted = preferred.strip()
    for name in workbook.sheetnames:
        if name.strip() == wanted:
            return name
    return None


def classification_to_legacy_mapped(classification: Any) -> dict[str, Any]:
    """Bridge trial balance classification to existing review/API mapped payload."""
    writes = build_cell_writes(classification)
    placements = writes_to_placements(writes)
    return {
        "generation_mode": "trial_balance",
        "exclude_sheets": list(EXCLUDED_SHEETS),
        "company_name": classification.company_name,
        "reporting_period": classification.period_label,
        "financial_year": classification.financial_year,
        "placements": placements,
        "trial_balance_validation": classification.totals.to_dict(),
        "line_items": classification.line_items,
    }
