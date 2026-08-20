"""End-to-end run against the project reference PDF and Excel sample.

Reads ``reference/ICAI_GN_Div_I_Sch_III.pdf`` and copies
``templates/Financial Statements_Sample.xlsx``. Neither original is modified.
"""

from __future__ import annotations

import json
import logging
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.classification.document_classifier import DocumentClassifier
from app.config import settings
from app.excel.template_comparator import compare_workbooks
from app.excel.template_manager import EXPECTED_SHEET_NAMES, canonical_sheet_name, file_sha256
from app.excel.workbook_generator import WorkbookGenerator
from app.extraction.pdf_extractor import PDFExtractor
from app.mapping.field_mapping import collect_source_items
from app.mapping.schedule_iii_mapper import ScheduleIIIMapper
from app.review.review_service import build_review
from app.validation.financial_validator import FinancialValidator

logger = logging.getLogger(__name__)

E2E_WORKBOOK_NAME = "test_generated_financial_statements.xlsx"
E2E_REPORT_NAME = "processing_report.json"


@dataclass
class EndToEndResult:
    """In-memory result of one reference-file pipeline run."""

    report: dict[str, Any]
    output_path: Path
    report_path: Path
    template_hash_before: str
    reference_hash_before: str
    template_hash_after: str
    reference_hash_after: str
    extraction_warnings: list[str]
    mapping_warnings: list[str]
    comparison_ok: bool
    comparison_errors: list[str]
    financial_validation: dict[str, Any]
    template_sheet_names: list[str]
    generated_sheet_names: list[str]
    written: list[dict[str, Any]] = field(default_factory=list)


def run_end_to_end(
    *,
    pdf_path: Path | None = None,
    template_path: Path | None = None,
    output_dir: Path | None = None,
) -> EndToEndResult:
    """Extract, classify, map, validate, and generate Excel from the reference files."""
    source_pdf = Path(pdf_path) if pdf_path else settings.reference_path
    template = Path(template_path) if template_path else settings.template_path
    destination_dir = Path(output_dir) if output_dir else settings.output_dir
    destination_dir.mkdir(parents=True, exist_ok=True)
    output_path = destination_dir / E2E_WORKBOOK_NAME
    report_path = destination_dir / E2E_REPORT_NAME

    if not source_pdf.exists():
        raise FileNotFoundError(f"Reference PDF was not found: {source_pdf}")
    if not template.exists():
        raise FileNotFoundError(f"Excel template was not found: {template}")
    if output_path.resolve() == template.resolve():
        raise ValueError("Refusing to generate onto the original template.")

    template_hash_before = file_sha256(template)
    reference_hash_before = file_sha256(source_pdf)
    warnings: list[str] = []
    errors: list[str] = []

    logger.info("E2E extract %s", source_pdf)
    extracted = PDFExtractor().extract(source_pdf, original_filename=source_pdf.name)
    extraction_warnings = [_issue_text(item) for item in (extracted.get("document") or {}).get("warnings") or []]
    extraction_warnings = [item for item in extraction_warnings if item]
    warnings.extend(extraction_warnings)
    pages_processed = int((extracted.get("document") or {}).get("pages") or 0)

    logger.info("E2E classify %s page(s)", pages_processed)
    classified_model = DocumentClassifier().classify(extracted)
    classified = classified_model.model_dump_classified()
    classification_warnings = [_issue_text(item) for item in classified.get("warnings") or []]
    warnings.extend(item for item in classification_warnings if item)

    logger.info("E2E map to Schedule III")
    mapped = ScheduleIIIMapper().map_classified(classified_model)
    mapping_warnings = [_issue_text(item) for item in mapped.get("warnings") or []]
    warnings.extend(item for item in mapping_warnings if item)

    sources = collect_source_items(classified)
    fields_extracted = len(sources)
    fields_mapped = sum(1 for item in (mapped.get("placements") or []) if item.get("action") == "write")

    review = build_review(mapped, classified)
    fields_requiring_review = int((review.get("summary") or {}).get("needs_review") or 0)

    logger.info("E2E financial validation")
    financial = FinancialValidator().validate(mapped, classified)
    validation_status = str(financial.get("status") or "WARNING")
    warnings.extend(_issue_text(item) for item in financial.get("warnings") or [])
    errors.extend(_issue_text(item) for item in financial.get("errors") or [])

    write_mapped = _mapped_without_uncertain_values(mapped, review)
    logger.info("E2E generate %s", output_path)
    generated = WorkbookGenerator(output_dir=destination_dir, template_path=template).generate_detailed(
        write_mapped,
        output_filename=E2E_WORKBOOK_NAME,
    )
    warnings.extend(_issue_text(item) for item in generated.validation.warnings)
    errors.extend(_issue_text(item) for item in generated.validation.errors)

    comparison = compare_workbooks(
        template,
        generated.path,
        allowed_value_cells=[
            (record["sheet"], record["cell"])
            for record in generated.written
            if record.get("sheet") and record.get("cell")
        ],
    )
    errors.extend(item.message for item in comparison.errors)
    warnings.extend(item.message for item in comparison.warnings)

    template_hash_after = file_sha256(template)
    reference_hash_after = file_sha256(source_pdf)
    if template_hash_after != template_hash_before:
        errors.append("The original Excel template was modified.")
    if reference_hash_after != reference_hash_before:
        errors.append("The original reference PDF was modified.")

    generated_book = load_workbook(generated.path, read_only=True)
    try:
        generated_sheet_names = list(generated_book.sheetnames)
    finally:
        generated_book.close()
    template_book = load_workbook(template, read_only=True)
    try:
        template_sheet_names = list(template_book.sheetnames)
    finally:
        template_book.close()

    report = {
        "input_file": str(source_pdf),
        "output_file": str(generated.path),
        "pages_processed": pages_processed,
        "fields_extracted": fields_extracted,
        "fields_mapped": fields_mapped,
        "fields_requiring_review": fields_requiring_review,
        "validation_status": validation_status,
        "warnings": _unique([item for item in warnings if item]),
        "errors": _unique([item for item in errors if item]),
    }
    _write_json(report_path, report)
    logger.info("E2E report written %s status=%s", report_path, validation_status)
    return EndToEndResult(
        report=report,
        output_path=generated.path,
        report_path=report_path,
        template_hash_before=template_hash_before,
        reference_hash_before=reference_hash_before,
        template_hash_after=template_hash_after,
        reference_hash_after=reference_hash_after,
        extraction_warnings=extraction_warnings,
        mapping_warnings=mapping_warnings,
        comparison_ok=comparison.ok,
        comparison_errors=[item.message for item in comparison.errors],
        financial_validation=financial,
        template_sheet_names=template_sheet_names,
        generated_sheet_names=generated_sheet_names,
        written=list(generated.written),
    )


def template_is_placeholder(sheet_names: list[str]) -> bool:
    return sheet_names == ["Placeholder"] or (
        len(sheet_names) == 1 and "placeholder" in sheet_names[0].lower()
    )


def expected_sheets_for(template_sheet_names: list[str]) -> list[str]:
    """Sheets the generated workbook must contain."""
    if template_is_placeholder(template_sheet_names):
        return list(template_sheet_names)
    present = {canonical_sheet_name(name) for name in template_sheet_names}
    missing_from_template = [
        name for name in EXPECTED_SHEET_NAMES if canonical_sheet_name(name) not in present
    ]
    if missing_from_template:
        return list(template_sheet_names)
    return list(template_sheet_names)


def _mapped_without_uncertain_values(mapped: dict[str, Any], review: dict[str, Any]) -> dict[str, Any]:
    """Keep high-confidence pending writes. Do not auto-write needs_review values."""
    allowed = {
        item.get("placement_index")
        for item in review.get("items") or []
        if item.get("status") == "pending" and item.get("action") == "write"
    }
    placements = [
        item
        for index, item in enumerate(mapped.get("placements") or [])
        if index in allowed
    ]
    filtered = dict(mapped)
    filtered["placements"] = placements
    return filtered


def _issue_text(item: Any) -> str:
    if item is None:
        return ""
    if isinstance(item, str):
        return item.strip()
    if isinstance(item, dict):
        return str(item.get("message") or item.get("detail") or item.get("error") or "").strip()
    message = getattr(item, "message", None)
    return str(message).strip() if message else str(item).strip()


def _unique(items: list[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for item in items:
        text = item.strip()
        if not text or text in seen:
            continue
        if "traceback" in text.lower():
            continue
        seen.add(text)
        result.append(text)
    return result


def _write_json(path: Path, payload: dict[str, Any]) -> None:
    temporary = path.with_name(f".{path.name}.tmp")
    temporary.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
    temporary.replace(path)


def main() -> int:
    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s [%(name)s] %(message)s")
    result = run_end_to_end()
    print(json.dumps(result.report, indent=2, ensure_ascii=False))
    return 0 if result.comparison_ok and result.template_hash_before == result.template_hash_after else 1


if __name__ == "__main__":
    raise SystemExit(main())
