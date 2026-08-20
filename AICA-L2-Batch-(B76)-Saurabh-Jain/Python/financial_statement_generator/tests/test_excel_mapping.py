"""Tests for configuration-driven Schedule III → Excel mapping."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.workbook.defined_name import DefinedName

from app.excel.template_manager import TemplateManager
from app.mapping.field_mapping import ExcelFieldMap, ExcelMappingEngine, load_excel_field_map


def _sourced(value, *, page: int = 1, confidence: float = 0.95, text: str = "") -> dict:
    return {
        "value": value,
        "source_page": page,
        "source_text": text or str(value),
        "confidence": confidence,
    }


def _line(label: str, current, previous=None, note=None) -> dict:
    return {
        "label": label,
        "mapping_code": None,
        "note_no": _sourced(note) if note is not None else _sourced(None, confidence=None, text=""),
        "current_period": _sourced(current, text=f"{label} {current}"),
        "previous_period": _sourced(previous, text=f"{label} {previous}") if previous is not None else _sourced(None, confidence=None, text=""),
    }


def _classified(*lines: dict, company_name: str | None = "ABC Private Limited") -> dict:
    payload = {
        "company": {
            "company_name": _sourced(company_name) if company_name else _sourced(None, confidence=None, text=""),
            "cin": _sourced(None, confidence=None, text=""),
        },
        "balance_sheet": {"line_items": list(lines), "identified": True},
        "profit_and_loss": {"line_items": []},
        "share_capital": {"line_items": list(lines)},
        "reserves_and_surplus": {"line_items": []},
        "warnings": [],
    }
    return payload


def _template(path: Path) -> dict:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "BS PnL"
    sheet["A12"] = "Particulars"
    sheet["C12"] = "Note"
    sheet["F12"] = "Current year"
    sheet["G12"] = "Previous year"
    sheet["A20"] = "Share capital"
    sheet["C20"] = None
    sheet["F20"] = None
    sheet["G20"] = None
    sheet["A21"] = "Reserves and surplus"
    sheet["F21"] = "=F20"
    sheet["G21"] = None
    sheet["A22"] = "Shareholders' funds"
    sheet["F22"] = "=F20+F21"
    sheet["A23"] = "Inventories"
    notes = workbook.create_sheet("Note 3 (Share Capital)")
    notes["A5"] = "Share capital"
    notes["B5"] = None
    workbook.defined_names.add(DefinedName(name="ShareCapitalCurrent", attr_text="'BS PnL'!$F$20"))
    workbook.save(path)
    workbook.close()
    return TemplateManager(path).inspect().to_dict()


def _engine(tmp_path: Path, field_map: ExcelFieldMap | None = None) -> ExcelMappingEngine:
    metadata = _template(tmp_path / "template.xlsx")
    return ExcelMappingEngine(
        field_map=field_map or load_excel_field_map(),
        icai_mappings={},
        template_metadata=metadata,
    )


def test_excel_field_map_loads_from_json() -> None:
    catalog = load_excel_field_map()
    assert catalog.fields["share_capital"].excel_sheet == "BS PnL"
    assert catalog.fields["share_capital"].note_sheet == "Note 3 (Share Capital)"
    assert catalog.fields["share_capital"].target_cells.current == "F13"
    assert catalog.fields["shareholders_funds_total"].kind == "derived_total"


def test_maps_by_template_label_not_configured_cell(tmp_path: Path) -> None:
    result = _engine(tmp_path).map(
        _classified(_line("Share capital", 100000, 80000, note="3"))
    )
    current = next(item for item in result["placements"] if item["field_key"] == "share_capital" and item["period"] == "current")
    previous = next(item for item in result["placements"] if item["field_key"] == "share_capital" and item["period"] == "previous")
    note = next(item for item in result["placements"] if item["field_key"] == "share_capital" and item["period"] == "note")
    assert current["excel_cell"] == "F20"
    assert previous["excel_cell"] == "G20"
    assert note["excel_cell"] == "C20"
    assert current["resolution"] == "template_label"
    assert current["extracted_value"] == 100000
    assert current["schedule_iii_category"] == "Shareholders' funds"
    assert current["action"] == "write"
    assert current["excel_cell"] != "F13"


def test_synonym_and_alias_labels_resolve(tmp_path: Path) -> None:
    result = _engine(tmp_path).map(_classified(_line("Equity share capital", 1250000, 1100000)))
    current = next(
        item
        for item in result["placements"]
        if item["field_key"] == "share_capital" and item["period"] == "current"
    )
    assert current["field_key"] == "share_capital"
    assert current["schedule_iii_label"] == "Share capital"
    assert current["extracted_value"] == 1250000


def test_skips_formula_cells_unless_override_requested(tmp_path: Path) -> None:
    result = _engine(tmp_path).map(
        _classified(
            _line("Share capital", 100000, 80000),
            _line("Reserves and surplus", 50000, 40000),
        )
    )
    reserves = next(
        item
        for item in result["placements"]
        if item["field_key"] == "reserves_and_surplus" and item["period"] == "current"
    )
    assert reserves["excel_cell"] == "F21"
    assert reserves["action"] == "skip_formula"
    assert reserves["extracted_value"] == 50000
    assert any("formula" in warning.lower() for warning in reserves["warnings"])


def test_derived_total_does_not_invent_missing_addends(tmp_path: Path) -> None:
    result = _engine(tmp_path).map(_classified(_line("Share capital", 100000, 80000)))
    derived = [item for item in result["placements"] if item["field_key"] == "shareholders_funds_total"]
    assert derived
    assert all(item["extracted_value"] is None for item in derived)
    assert all(item["action"] == "missing_value" for item in derived)
    assert any("addends" in warning for item in derived for warning in item["warnings"])


def test_derived_total_sums_when_addends_exist_and_preserves_formula(tmp_path: Path) -> None:
    result = _engine(tmp_path).map(
        _classified(
            _line("Share capital", 100000, 80000),
            _line("Reserves and surplus", 50000, 40000),
            _line("Money received against share warrants", 0, 0),
        )
    )
    derived = next(
        item
        for item in result["placements"]
        if item["field_key"] == "shareholders_funds_total" and item["period"] == "current"
    )
    assert derived["extracted_value"] == 150000
    assert derived["action"] == "skip_formula"


def test_does_not_invent_unmapped_or_missing_values(tmp_path: Path) -> None:
    result = _engine(tmp_path).map(
        _classified(_line("Share capital", 100000, 80000), _line("Mystery caption", 9))
    )
    keys = {item["field_key"] for item in result["placements"]}
    assert "inventories" not in keys
    assert any(item["source_label"] == "Mystery caption" for item in result["unmapped_sources"])
    written = [item for item in result["placements"] if item["action"] == "write" and item["field_key"] == "inventories"]
    assert written == []


def test_mapping_report_has_required_columns(tmp_path: Path) -> None:
    result = _engine(tmp_path).map(_classified(_line("Share capital", 100000, 80000, note="3")))
    rows = result["report"]["rows"]
    assert rows
    required = {
        "source_label",
        "extracted_value",
        "schedule_iii_category",
        "excel_destination",
        "confidence",
        "warnings",
    }
    for row in rows:
        assert required <= set(row)
    summary = result["report"]["summary"]
    assert summary["mapped"] >= 1
    assert "skipped_formulas" in summary


def test_defined_name_resolution_beats_coordinates(tmp_path: Path) -> None:
    catalog = load_excel_field_map()
    catalog.fields["share_capital"].defined_names.current = "ShareCapitalCurrent"
    result = _engine(tmp_path, field_map=catalog).map(_classified(_line("Share capital", 42, 7)))
    current = next(item for item in result["placements"] if item["field_key"] == "share_capital" and item["period"] == "current")
    assert current["excel_cell"] == "F20"
    assert current["resolution"] == "defined_name"
    assert current["extracted_value"] == 42


def test_resolves_sample_sheet_names_with_trailing_spaces() -> None:
    catalog = load_excel_field_map()
    available = ["BS PnL", "Cash Flow ", "Stock Reconciliation "]
    assert catalog.resolve_sheet_name("Cash Flow", available) == "Cash Flow "
    assert catalog.resolve_sheet_name("Stock Reconciliation", available) == "Stock Reconciliation "


def test_discovers_previous_column_from_second_as_at_header(tmp_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "BS PnL"
    sheet["A8"] = "Particulars"
    sheet["F8"] = "As at March 31, 2025"
    sheet["G8"] = "As at March 31, 2024"
    sheet["A13"] = "Share capital"
    sheet["F13"] = None
    sheet["G13"] = None
    path = tmp_path / "years.xlsx"
    workbook.save(path)
    workbook.close()
    engine = ExcelMappingEngine(
        field_map=load_excel_field_map(),
        icai_mappings={},
        template_metadata=TemplateManager(path).inspect().to_dict(),
    )
    result = engine.map(_classified(_line("Share capital", 100000, 80000)))
    current = next(item for item in result["placements"] if item["field_key"] == "share_capital" and item["period"] == "current")
    previous = next(item for item in result["placements"] if item["field_key"] == "share_capital" and item["period"] == "previous")
    assert current["excel_cell"] == "F13"
    assert previous["excel_cell"] == "G13"


def test_configured_cell_used_only_when_label_missing(tmp_path: Path) -> None:
    metadata = _template(tmp_path / "template.xlsx")
    bs = next(sheet for sheet in metadata["sheets"] if sheet["name"] == "BS PnL")
    bs["cells"].pop("A20", None)
    engine = ExcelMappingEngine(field_map=load_excel_field_map(), icai_mappings={}, template_metadata=metadata)
    result = engine.map(_classified(_line("Share capital", 100000, 80000)))
    current = next(item for item in result["placements"] if item["field_key"] == "share_capital" and item["period"] == "current")
    assert current["excel_cell"] == "F13"
    assert current["resolution"] == "configured_cell"
    assert any("configured cell" in warning.lower() for warning in current["warnings"])
