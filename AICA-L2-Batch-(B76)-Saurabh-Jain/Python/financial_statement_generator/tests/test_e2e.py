"""End-to-end pipeline against the project reference PDF and Excel sample."""

from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.config import settings
from app.e2e import expected_sheets_for, run_end_to_end, template_is_placeholder
from app.excel.template_comparator import compare_workbooks
from app.excel.template_manager import EXPECTED_SHEET_NAMES, canonical_sheet_name, file_sha256


@pytest.fixture(scope="module")
def e2e():
    if not settings.reference_path.exists():
        pytest.skip(f"Reference PDF missing: {settings.reference_path}")
    if not settings.template_path.exists():
        pytest.skip(f"Excel template missing: {settings.template_path}")
    return run_end_to_end(
        pdf_path=settings.reference_path,
        template_path=settings.template_path,
        output_dir=settings.output_dir,
    )


def test_original_reference_files_are_not_modified(e2e) -> None:
    assert e2e.template_hash_before == e2e.template_hash_after
    assert e2e.reference_hash_before == e2e.reference_hash_after
    assert file_sha256(settings.template_path) == e2e.template_hash_before
    assert file_sha256(settings.reference_path) == e2e.reference_hash_before
    assert e2e.output_path.resolve() != settings.template_path.resolve()
    assert e2e.output_path.resolve() != settings.reference_path.resolve()


def test_generated_workbook_opens_successfully(e2e) -> None:
    assert e2e.output_path.exists()
    assert e2e.output_path.name == "test_generated_financial_statements.xlsx"
    workbook = load_workbook(e2e.output_path)
    try:
        assert workbook.sheetnames
        assert workbook.worksheets
        first = workbook.worksheets[0]
        assert first.max_row >= 1
    finally:
        workbook.close()


def test_expected_sheets_exist_and_order_is_preserved(e2e) -> None:
    template_names = e2e.template_sheet_names
    generated_names = e2e.generated_sheet_names
    expected = expected_sheets_for(template_names)
    missing = [name for name in expected if name not in generated_names]
    assert not missing, f"Generated workbook is missing expected sheets: {missing}"
    assert generated_names == template_names
    extra = [name for name in generated_names if name not in template_names]
    deleted = [name for name in template_names if name not in generated_names]
    assert extra == [], f"Unexpected sheets were created: {extra}"
    assert deleted == [], f"Unexpected sheets were deleted: {deleted}"
    if not template_is_placeholder(template_names):
        generated_keys = {canonical_sheet_name(name) for name in generated_names}
        still_missing = [
            name for name in EXPECTED_SHEET_NAMES if canonical_sheet_name(name) not in generated_keys
        ]
        assert not still_missing, f"ICAI sample sheets missing: {still_missing}"


def test_layout_formulas_and_formatting_are_preserved(e2e) -> None:
    report = compare_workbooks(
        settings.template_path,
        e2e.output_path,
        allowed_value_cells=[
            (record["sheet"], record["cell"])
            for record in e2e.written
            if record.get("sheet") and record.get("cell")
        ],
    )
    layout_categories = {
        "merged_cells",
        "font",
        "font_size",
        "bold",
        "italic",
        "borders",
        "alignment",
        "number_format",
        "formula",
        "sheet_order",
        "sheet_name",
        "row_heights",
        "column_widths",
        "freeze_panes",
        "print_area",
        "header_footer",
        "hidden_rows",
        "hidden_columns",
        "page_orientation",
    }
    layout_errors = [item.message for item in report.errors if item.category in layout_categories]
    assert layout_errors == [], layout_errors


def test_financial_validation_runs(e2e) -> None:
    validation = e2e.financial_validation
    assert validation["status"] in {"PASS", "WARNING", "ERROR"}
    assert "checks" in validation
    assert isinstance(validation["checks"], list)
    assert e2e.report["validation_status"] == validation["status"]


def test_extraction_and_mapping_warnings_are_reported(e2e) -> None:
    assert isinstance(e2e.extraction_warnings, list)
    assert isinstance(e2e.mapping_warnings, list)
    reported = e2e.report["warnings"]
    for warning in e2e.extraction_warnings:
        assert warning in reported
    for warning in e2e.mapping_warnings:
        assert warning in reported


def test_processing_report_is_written(e2e) -> None:
    path = Path(settings.output_dir) / "processing_report.json"
    assert path.exists()
    payload = json.loads(path.read_text(encoding="utf-8"))
    assert payload["input_file"] == str(settings.reference_path)
    assert payload["output_file"].endswith("test_generated_financial_statements.xlsx")
    assert payload["pages_processed"] >= 1
    assert payload["fields_extracted"] >= 0
    assert payload["fields_mapped"] >= 0
    assert payload["fields_requiring_review"] >= 0
    assert payload["validation_status"] in {"PASS", "WARNING", "ERROR"}
    assert isinstance(payload["warnings"], list)
    assert isinstance(payload["errors"], list)
    assert "traceback" not in json.dumps(payload).lower()
    assert payload == e2e.report
