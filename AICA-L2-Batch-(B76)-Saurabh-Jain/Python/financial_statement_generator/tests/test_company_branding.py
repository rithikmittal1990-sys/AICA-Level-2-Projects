"""Tests for replacing ICAI sample company branding."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from app.excel.company_branding import apply_company_branding, company_name_from_placements
from app.excel.workbook_generator import WorkbookGenerator
from app.mapping.trial_balance_placements import prepare_trial_balance_mapped
from app.mapping.schedule_iii_mapper import ScheduleIIIMapper
from app.classification.trial_balance_classifier import TrialBalanceClassifier
from app.extraction.trial_balance_reader import TrialBalanceReader
from tests.test_trial_balance import _sample_trial_balance


def test_company_name_from_placements() -> None:
    name = company_name_from_placements(
        [{"field_key": "company_name", "extracted_value": "XYZ Private Limited"}]
    )
    assert name == "XYZ Private Limited"


def test_apply_company_branding_replaces_sample_name() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "BS PnL"
    sheet["A2"] = "ABC INDIA LIMITED"
    sheet["A58"] = "For ABC & Company"
    sheet["E60"] = "ABC INDIA LIMITED"
    cash = workbook.create_sheet("Cash Flow ")
    cash["B2"] = "ABC INDIA LIMITED\n"
    written = apply_company_branding(
        workbook,
        "XYZ Private Limited",
        template_names=("ABC INDIA LIMITED", "ABC & Company"),
        full_workbook=True,
    )
    assert sheet["A2"].value == "XYZ Private Limited"
    assert sheet["E60"].value == "XYZ Private Limited"
    assert "ABC" not in str(sheet["A58"].value)
    assert "XYZ Private Limited" in str(cash["B2"].value)
    assert len(written) >= 3
    workbook.close()


def test_trial_balance_generation_uses_uploaded_company_and_amounts(tmp_path: Path) -> None:
    path = _sample_trial_balance(tmp_path / "TrialBal.xlsx")
    document = TrialBalanceReader(upload_dir=tmp_path).read_path(path)
    classified = TrialBalanceClassifier().classify(document)
    classified_dict = classified.model_dump_classified()
    mapped = ScheduleIIIMapper().map_classified(classified)
    mapped = prepare_trial_balance_mapped(mapped, classified_dict)
    for placement in mapped["placements"]:
        placement["review_status"] = "approved"
    result = WorkbookGenerator(output_dir=tmp_path / "out").generate_detailed(mapped)
    workbook = load_workbook(result.path)
    try:
        assert "Cash Flow" not in {name.strip() for name in workbook.sheetnames}
        sheet = workbook["BS PnL"]
        assert sheet["A2"].value == "XYZ Private Limited"
        assert sheet["E60"].value == "XYZ Private Limited"
        share = next(
            item
            for item in mapped["placements"]
            if item.get("field_key") == "share_capital" and item.get("excel_cell") == "F13"
        )
        assert sheet[share["excel_cell"]].value == 1000000
        revenue = next(
            item
            for item in mapped["placements"]
            if item.get("field_key") == "revenue_from_operations" and item.get("excel_sheet") == "Note 20-31"
        )
        notes = workbook["Note 20-31"]
        assert notes[revenue["excel_cell"]].value == 618478
    finally:
        workbook.close()
