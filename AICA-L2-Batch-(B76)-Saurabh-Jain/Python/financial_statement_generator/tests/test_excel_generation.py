"""Tests for Excel template inspection and copying."""

from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.config import settings
from app.excel.formatting import apply_standard_formatting, write_text, write_value
from app.excel.template_manager import (
    EXPECTED_SHEET_NAMES,
    TemplateManager,
    canonical_sheet_name,
    file_sha256,
)
from app.excel.workbook_generator import (
    DEFAULT_OUTPUT_FILENAME,
    WorkbookGenerator,
    copy_template,
    validate_workbook,
)

REQUIRED_SHEET_METADATA_KEYS = {
    "name",
    "index",
    "order",
    "merged_cells",
    "row_heights",
    "column_widths",
    "freeze_panes",
    "print_area",
    "page_setup",
    "formulas",
    "cells",
    "hidden_rows",
    "hidden_columns",
    "sheet_state",
    "blank_layout_cells",
}

REQUIRED_CELL_METADATA_KEYS = {
    "value",
    "formula",
    "number_format",
    "font",
    "font_size",
    "bold",
    "italic",
    "alignment",
    "border",
    "blank",
}


def test_excel_module_importable() -> None:
    assert apply_standard_formatting is not None
    assert TemplateManager is not None
    assert WorkbookGenerator is not None


def test_template_file_exists() -> None:
    assert settings.template_path.exists()
    assert settings.template_path.suffix.lower() == ".xlsx"


def test_template_loads_successfully() -> None:
    manager = TemplateManager()
    workbook = manager.load()
    try:
        assert workbook.sheetnames
        assert workbook.worksheets
        first = workbook.worksheets[0]
        assert first.max_row >= 1
        assert first.max_column >= 1
    finally:
        workbook.close()


def test_inspect_does_not_modify_original_template() -> None:
    manager = TemplateManager()
    before = file_sha256(manager.template_path)
    metadata = manager.inspect()
    after = file_sha256(manager.template_path)
    assert before == after
    assert metadata.source_sha256 == before


def test_metadata_captures_required_layout_fields() -> None:
    metadata = TemplateManager().inspect().to_dict()
    assert metadata["sheet_names"]
    assert metadata["sheet_order"] == metadata["sheet_names"]
    assert metadata["expected_sheet_names"] == list(EXPECTED_SHEET_NAMES)
    assert "workbook_formulas" in metadata
    assert "defined_names" in metadata

    for sheet in metadata["sheets"]:
        missing = REQUIRED_SHEET_METADATA_KEYS - sheet.keys()
        assert not missing, f"{sheet['name']} missing {sorted(missing)}"
        for cell in sheet["cells"].values():
            missing_cell = REQUIRED_CELL_METADATA_KEYS - cell.keys()
            assert not missing_cell


def test_export_metadata_json(tmp_path: Path) -> None:
    destination = tmp_path / "template_metadata.json"
    written = TemplateManager().export_metadata(destination)
    assert written.exists()
    payload = json.loads(written.read_text(encoding="utf-8"))
    assert payload["sheets"]
    assert payload["sheet_order"]


def test_copy_template_preserves_bytes_and_leaves_original(tmp_path: Path) -> None:
    manager = TemplateManager()
    original_hash = file_sha256(manager.template_path)
    copy_path = manager.copy_template(tmp_path / "working_copy.xlsx")
    assert copy_path.exists()
    assert copy_path.resolve() != manager.template_path.resolve()
    assert file_sha256(copy_path) == original_hash
    assert file_sha256(manager.template_path) == original_hash

    copied = load_workbook(copy_path)
    original = manager.load()
    try:
        assert copied.sheetnames == original.sheetnames
    finally:
        copied.close()
        original.close()


def test_copy_template_refuses_to_overwrite_original() -> None:
    manager = TemplateManager()
    with pytest.raises(ValueError, match="itself"):
        manager.copy_template(manager.template_path)


def test_expected_sheet_names_are_documented() -> None:
    assert "BS PnL" in EXPECTED_SHEET_NAMES
    assert "Cash Flow" in EXPECTED_SHEET_NAMES
    assert "Note 3 (Share Capital)" in EXPECTED_SHEET_NAMES
    assert "Note 12 (PPE)" in EXPECTED_SHEET_NAMES
    assert "13.CWIP" in EXPECTED_SHEET_NAMES
    assert "Ratio_working" in EXPECTED_SHEET_NAMES
    assert "EPS" in EXPECTED_SHEET_NAMES
    assert len(EXPECTED_SHEET_NAMES) == 15


def test_inspect_captures_layout_features_on_sample_workbook(tmp_path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, Side
    from openpyxl.workbook.defined_name import DefinedName

    path = tmp_path / "layout_sample.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "BS PnL"
    sheet["A1"] = "Particulars"
    sheet["A1"].font = Font(name="Times New Roman", size=14, bold=True, italic=False)
    sheet["A1"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    sheet["B1"] = 1000
    sheet["B1"].number_format = "#,##0.00"
    sheet["B1"].border = Border(bottom=Side(style="thin"))
    sheet.merge_cells("A2:B2")
    sheet["A2"] = "Merged heading"
    sheet["C1"] = "=B1*2"
    sheet["D5"] = None
    sheet["D5"].border = Border(left=Side(style="thin"))
    sheet.row_dimensions[1].height = 22
    sheet.column_dimensions["A"].width = 28
    sheet.row_dimensions[8].hidden = True
    sheet.column_dimensions["Z"].hidden = True
    sheet.freeze_panes = "A2"
    sheet.print_area = "A1:C10"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToPage = True
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 1
    workbook.create_sheet("Cash Flow")
    workbook.defined_names.add(DefinedName(name="TotalIncome", attr_text="'BS PnL'!$B$1"))
    workbook.save(path)
    workbook.close()

    metadata = TemplateManager(path).inspect().to_dict()
    first = metadata["sheets"][0]
    assert first["name"] == "BS PnL"
    assert first["order"] == 0
    assert "A2:B2" in first["merged_cells"]
    assert first["row_heights"]["1"]["height"] == 22
    assert first["column_widths"]["A"]["width"] == 28
    assert first["cells"]["A1"]["font"]["name"] == "Times New Roman"
    assert first["cells"]["A1"]["font_size"] == 14
    assert first["cells"]["A1"]["bold"] is True
    assert first["cells"]["A1"]["italic"] is False
    assert first["cells"]["A1"]["alignment"]["wrap_text"] is True
    assert first["cells"]["B1"]["border"]["bottom"]["style"] == "thin"
    assert first["cells"]["B1"]["number_format"] == "#,##0.00"
    assert first["formulas"]["C1"] == "=B1*2"
    assert first["cells"]["C1"]["formula"] == "=B1*2"
    assert "D5" in first["blank_layout_cells"]
    assert 8 in first["hidden_rows"]
    assert "Z" in first["hidden_columns"]
    assert first["freeze_panes"] == "A2"
    assert "A1:C10" in str(first["print_area"]).replace("$", "")
    assert first["page_setup"]["orientation"] == "landscape"
    assert metadata["sheet_order"] == ["BS PnL", "Cash Flow"]
    assert any(item["name"] == "TotalIncome" for item in metadata["defined_names"])
    assert file_sha256(path) == metadata["source_sha256"]


def test_expected_sheets_present_when_sample_is_complete() -> None:
    manager = TemplateManager()
    workbook = manager.load()
    try:
        names = workbook.sheetnames
    finally:
        workbook.close()

    if names == ["Placeholder"] or (len(names) == 1 and "placeholder" in names[0].lower()):
        pytest.skip(
            "Place the ICAI Division I sample workbook at "
            "templates/Financial Statements_Sample.xlsx to validate sheet names."
        )

    present = {canonical_sheet_name(name) for name in names}
    missing = [name for name in EXPECTED_SHEET_NAMES if canonical_sheet_name(name) not in present]
    assert not missing, f"Template is missing expected sheets: {missing}"


def _styled_sample(path: Path) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, Side
    from openpyxl.workbook.defined_name import DefinedName

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "BS PnL"
    sheet["A1"] = "Particulars"
    sheet["A1"].font = Font(name="Times New Roman", size=14, bold=True)
    sheet["A1"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    sheet["B1"] = 1000
    sheet["B1"].number_format = "#,##0.00"
    sheet["B1"].border = Border(bottom=Side(style="thin"))
    sheet.merge_cells("A2:B2")
    sheet["A2"] = "Merged heading"
    sheet["C1"] = "=B1*2"
    sheet["D5"] = None
    sheet["D5"].border = Border(left=Side(style="thin"))
    sheet.row_dimensions[1].height = 22
    sheet.column_dimensions["A"].width = 28
    sheet.row_dimensions[8].hidden = True
    sheet.column_dimensions["Z"].hidden = True
    sheet.freeze_panes = "A2"
    sheet.print_area = "A1:C10"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToPage = True
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 1
    workbook.create_sheet("Cash Flow")
    workbook.defined_names.add(DefinedName(name="TotalIncome", attr_text="'BS PnL'!$B$1"))
    workbook.save(path)
    workbook.close()
    return path


def test_write_value_preserves_formatting(tmp_path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, Side

    path = tmp_path / "cell.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet["B2"] = 10
    sheet["B2"].font = Font(name="Times New Roman", size=12, bold=True)
    sheet["B2"].number_format = "#,##0.00"
    sheet["B2"].alignment = Alignment(horizontal="right")
    sheet["B2"].border = Border(bottom=Side(style="thin"))
    sheet["C2"] = "=B2*3"
    workbook.save(path)
    workbook.close()

    workbook = load_workbook(path)
    sheet = workbook.active
    written = write_value(sheet, "B2", 2500)
    skipped = write_value(sheet, "C2", 99)
    assert written.written is True
    assert skipped.skipped is True
    assert skipped.reason == "formula"
    assert sheet["B2"].value == 2500
    assert sheet["B2"].font.name == "Times New Roman"
    assert sheet["B2"].font.bold is True
    assert sheet["B2"].number_format == "#,##0.00"
    assert sheet["B2"].alignment.horizontal == "right"
    assert sheet["B2"].border.bottom.style == "thin"
    assert sheet["C2"].value == "=B2*3"
    write_text(sheet, "A1", "Share capital")
    assert sheet["A1"].value == "Share capital"
    workbook.close()


def test_copy_template_never_writes_to_original() -> None:
    original = file_sha256(settings.template_path)
    copied = copy_template()
    try:
        assert copied.name == DEFAULT_OUTPUT_FILENAME
        assert copied.resolve() != settings.template_path.resolve()
        assert file_sha256(settings.template_path) == original
        assert file_sha256(copied) == original
    finally:
        if copied.exists() and copied.resolve() != settings.template_path.resolve():
            copied.unlink()


def test_generator_copies_template_and_preserves_structure(tmp_path: Path) -> None:
    template = _styled_sample(tmp_path / "Financial Statements_Sample.xlsx")
    original_hash = file_sha256(template)
    generator = WorkbookGenerator(output_dir=tmp_path / "out", template_path=template)
    result = generator.generate_detailed(
        {
            "placements": [
                {
                    "field_key": "share_capital",
                    "excel_sheet": "BS PnL",
                    "excel_cell": "B1",
                    "extracted_value": 2500,
                    "action": "write",
                    "period": "current",
                },
                {
                    "field_key": "tax_expense",
                    "excel_sheet": "BS PnL",
                    "excel_cell": "C1",
                    "extracted_value": 999,
                    "action": "write",
                    "period": "current",
                },
                {
                    "field_key": "company_name",
                    "excel_sheet": "BS PnL",
                    "excel_cell": "A1",
                    "extracted_value": "ABC Private Limited",
                    "action": "write",
                    "value_role": "text",
                },
            ]
        }
    )

    assert result.path.name == DEFAULT_OUTPUT_FILENAME
    assert result.path.resolve() != template.resolve()
    assert file_sha256(template) == original_hash
    assert result.validation.ok, result.validation.errors

    generated = load_workbook(result.path)
    original = load_workbook(template)
    try:
        sheet = generated["BS PnL"]
        assert generated.sheetnames == original.sheetnames
        assert sheet["B1"].value == 2500
        assert sheet["B1"].number_format == "#,##0.00"
        assert sheet["B1"].font.name == "Times New Roman" or sheet["B1"].number_format == "#,##0.00"
        assert sheet["B1"].border.bottom.style == "thin"
        assert sheet["C1"].value == "=B1*2"
        assert sheet["A1"].value == "ABC Private Limited"
        assert sheet["A1"].font.name == "Times New Roman"
        assert sheet["A1"].font.bold is True
        assert sheet["A1"].font.size == 14
        assert "A2:B2" in {str(item) for item in sheet.merged_cells.ranges}
        assert sheet.row_dimensions[1].height == 22
        assert sheet.column_dimensions["A"].width == 28
        assert sheet.row_dimensions[8].hidden is True
        assert sheet.column_dimensions["Z"].hidden is True
        assert sheet.freeze_panes == "A2"
        assert "A1:C10" in str(sheet.print_area).replace("$", "")
        assert sheet.page_setup.orientation == "landscape"
        skipped_formula = next(item for item in result.skipped if item["cell"] == "C1" or item["reason"] == "formula")
        assert skipped_formula["written"] is False
    finally:
        generated.close()
        original.close()


def test_validate_workbook_detects_sheet_order_change(tmp_path: Path) -> None:
    template = _styled_sample(tmp_path / "sample.xlsx")
    generated = tmp_path / "broken.xlsx"
    copy_template(generated, template_path=template)
    workbook = load_workbook(generated)
    workbook.move_sheet("Cash Flow", offset=-1)
    workbook.save(generated)
    workbook.close()
    report = validate_workbook(generated, template)
    assert report.ok is False
    assert any("Sheet order" in error for error in report.errors)


def test_generator_resolves_trailing_space_sheet_names(tmp_path: Path) -> None:
    from openpyxl import Workbook

    template = tmp_path / "sample.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Cash Flow "
    sheet["D11"] = None
    workbook.save(template)
    workbook.close()

    generator = WorkbookGenerator(output_dir=tmp_path / "out", template_path=template)
    result = generator.generate_detailed(
        {
            "placements": [
                {
                    "field_key": "cash_flow_from_operating_activities",
                    "excel_sheet": "Cash Flow",
                    "excel_cell": "D11",
                    "extracted_value": 1234,
                    "action": "write",
                    "period": "current",
                }
            ]
        }
    )
    assert result.validation.ok, result.validation.errors
    written = next(item for item in result.written if item["cell"] == "D11")
    assert written["sheet"] == "Cash Flow "
    generated = load_workbook(result.path)
    try:
        assert generated["Cash Flow "]["D11"].value == 1234
    finally:
        generated.close()


def test_placeholder_template_generation_leaves_original_untouched(tmp_path: Path) -> None:
    original_hash = file_sha256(settings.template_path)
    generator = WorkbookGenerator(output_dir=tmp_path)
    path = generator.generate({}, DEFAULT_OUTPUT_FILENAME)
    assert path.exists()
    assert path.name == DEFAULT_OUTPUT_FILENAME
    assert file_sha256(settings.template_path) == original_hash
    assert generator.last_result is not None
    assert generator.last_result.validation.ok, generator.last_result.validation.errors


def test_apply_standard_formatting_does_not_restyle_sheet(tmp_path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "Share capital"
    sheet["A1"].font = Font(name="Times New Roman", size=14, bold=True)
    apply_standard_formatting(sheet)
    assert sheet["A1"].font.name == "Times New Roman"
    assert sheet["A1"].font.size == 14
    workbook.close()

