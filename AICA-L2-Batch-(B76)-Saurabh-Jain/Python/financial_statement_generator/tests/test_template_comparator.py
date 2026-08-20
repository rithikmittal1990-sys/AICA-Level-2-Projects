"""Tests for strict template vs generated workbook comparison."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, Side

from app.excel.template_comparator import compare_workbooks
from app.excel.template_manager import file_sha256
from app.excel.workbook_generator import copy_template, validate_workbook


def _sample(path: Path) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "BS PnL"
    sheet["A1"] = "Particulars"
    sheet["A1"].font = Font(name="Times New Roman", size=14, bold=True, italic=False)
    sheet["A1"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    sheet["B1"] = 1000
    sheet["B1"].number_format = "#,##0.00"
    sheet["B1"].border = Border(bottom=Side(style="thin"))
    sheet["C1"] = "=B1*2"
    sheet.merge_cells("A2:B2")
    sheet["A2"] = "Merged heading"
    sheet.row_dimensions[1].height = 22
    sheet.column_dimensions["A"].width = 28
    sheet.row_dimensions[8].hidden = True
    sheet.column_dimensions["Z"].hidden = True
    sheet.freeze_panes = "A2"
    sheet.print_area = "A1:C10"
    sheet.page_setup.orientation = "landscape"
    sheet.oddHeader.center.text = "Division I"
    sheet.oddFooter.left.text = "Confidential"
    workbook.create_sheet("Cash Flow")
    workbook.save(path)
    workbook.close()
    return path


def test_identical_copy_has_no_layout_errors(tmp_path: Path) -> None:
    template = _sample(tmp_path / "Financial Statements_Sample.xlsx")
    generated = copy_template(tmp_path / "Financial_Statements_Generated.xlsx", template_path=template)
    report = compare_workbooks(template, generated)
    assert report.ok, [item.message for item in report.errors]
    assert report.errors == []


def test_numeric_value_change_is_ignored(tmp_path: Path) -> None:
    template = _sample(tmp_path / "sample.xlsx")
    generated = copy_template(tmp_path / "generated.xlsx", template_path=template)
    workbook = load_workbook(generated)
    workbook["BS PnL"]["B1"] = 2500
    workbook.save(generated)
    workbook.close()

    report = compare_workbooks(template, generated)
    assert report.ok, [item.message for item in report.errors]
    assert any("expected value change" in item.message for item in report.ignored)


def test_label_change_is_an_error_unless_allowed(tmp_path: Path) -> None:
    template = _sample(tmp_path / "sample.xlsx")
    generated = copy_template(tmp_path / "generated.xlsx", template_path=template)
    workbook = load_workbook(generated)
    workbook["BS PnL"]["A1"] = "Changed heading"
    workbook.save(generated)
    workbook.close()

    report = compare_workbooks(template, generated)
    assert report.ok is False
    assert any("unexpected value change" in item.message for item in report.errors)

    allowed = compare_workbooks(template, generated, allowed_value_cells=[("BS PnL", "A1")])
    assert allowed.ok, [item.message for item in allowed.errors]


def test_italic_and_header_changes_are_reported(tmp_path: Path) -> None:
    template = _sample(tmp_path / "sample.xlsx")
    generated = copy_template(tmp_path / "generated.xlsx", template_path=template)
    workbook = load_workbook(generated)
    sheet = workbook["BS PnL"]
    sheet["A1"].font = Font(name="Times New Roman", size=14, bold=True, italic=True)
    sheet.oddHeader.center.text = "Redesigned"
    workbook.save(generated)
    workbook.close()

    report = compare_workbooks(template, generated)
    messages = [item.message for item in report.errors]
    assert any("italic" in message for message in messages)
    assert any("header" in message for message in messages)


def test_formula_and_merged_cells_must_be_preserved(tmp_path: Path) -> None:
    template = _sample(tmp_path / "sample.xlsx")
    generated = copy_template(tmp_path / "generated.xlsx", template_path=template)
    workbook = load_workbook(generated)
    sheet = workbook["BS PnL"]
    sheet["C1"] = 99
    sheet.unmerge_cells("A2:B2")
    workbook.save(generated)
    workbook.close()

    report = compare_workbooks(template, generated)
    messages = [item.message for item in report.errors]
    assert any("formula" in message for message in messages)
    assert any("merged cells" in message for message in messages)


def test_validate_workbook_uses_comparator(tmp_path: Path) -> None:
    template = _sample(tmp_path / "sample.xlsx")
    generated = copy_template(tmp_path / "generated.xlsx", template_path=template)
    workbook = load_workbook(generated)
    workbook.move_sheet("Cash Flow", offset=-1)
    workbook.save(generated)
    workbook.close()
    report = validate_workbook(generated, template)
    assert report.ok is False
    assert any("Sheet order" in error for error in report.errors)


def test_comparison_does_not_modify_files(tmp_path: Path) -> None:
    template = _sample(tmp_path / "sample.xlsx")
    generated = copy_template(tmp_path / "generated.xlsx", template_path=template)
    before_template = file_sha256(template)
    before_generated = file_sha256(generated)
    compare_workbooks(template, generated)
    assert file_sha256(template) == before_template
    assert file_sha256(generated) == before_generated
